import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo

def gerar_tabelas_e_pivots(pasta_origem, pasta_destino, coluna_filtro, valores_desejados, colunas_indice, colunas_valores, coluna_operacao):
    """Gera tabelas e tabelas dinâmicas separadas por material a partir de arquivos Excel."""
    try:
        # Cria a pasta de destino se não existir
        if not os.path.exists(pasta_destino):
            os.makedirs(pasta_destino)
            print(f"Pasta de destino criada: {pasta_destino}")

        # Lista arquivos Excel (xlsx, xls, xlsm, xlsb)
        arquivos = [os.path.join(pasta_origem, arquivo) for arquivo in os.listdir(pasta_origem) if arquivo.endswith(('.xlsx', '.xls', '.xlsm', '.xlsb', '.XLSX'))]
        if not arquivos:
            print(f"Nenhum arquivo Excel encontrado na pasta de origem: {pasta_origem}")
            return # Sai da função se não houver arquivos

        for arquivo in arquivos:
            print(f"\nProcessando arquivo: {arquivo}")
            try:
                df = pd.read_excel(arquivo, index_col=None)
                print(f"Arquivo '{arquivo}' lido com sucesso. Shape: {df.shape}")

                if coluna_filtro not in df.columns:
                    print(f"A coluna de filtro '{coluna_filtro}' não existe no arquivo {arquivo}. Pulando este arquivo.")
                    continue

                df_filtrado = df[df[coluna_filtro].isin(valores_desejados)]
                print(f"Tamanho do DataFrame filtrado: {len(df_filtrado)}")

                if df_filtrado.empty:
                    print(f"Nenhum dado corresponde aos critérios de filtro no arquivo {arquivo}. Pulando este arquivo.")
                    continue

                if 'Material' not in df_filtrado.columns:
                    print(f"A coluna 'Material' não existe no arquivo {arquivo}. Pulando este arquivo.")
                    continue

                for material in df_filtrado['Material'].unique():
                    print(f"\nProcessando material: {material}")
                    material_df = df_filtrado[df_filtrado['Material'] == material]

                    if material_df.empty:
                        print(f"Nenhum dado encontrado para o material {material}. Pulando.")
                        continue

                    nome_arquivo_material = f"BD_103_{material}.xlsx"
                    caminho_arquivo_material = os.path.join(pasta_destino, nome_arquivo_material)
                    print(f"Criando arquivo: {caminho_arquivo_material}")

                    try:
                        with pd.ExcelWriter(caminho_arquivo_material, engine="openpyxl") as writer:
                            material_df.to_excel(writer, sheet_name="Dados", index=False)

                            if not all(col in material_df.columns for col in colunas_indice + colunas_valores + [coluna_operacao]):
                                print(f"Colunas necessárias para o pivot não encontradas para o material {material}. Pulando a criação do Pivot.")
                                continue

                            pivot_df = pd.pivot_table(
                                material_df,
                                values=colunas_valores,
                                index=colunas_indice,
                                columns=coluna_operacao,
                                aggfunc="sum",
                                fill_value=0
                            ).reset_index()

                            if pivot_df.empty:
                                print(f"Tabela Pivot vazia para o material {material}. Pulando formatação.")
                                continue

                            pivot_df.columns = [''.join(map(str, col)).strip() for col in pivot_df.columns.values]
                            valores_pivot = [col for col in pivot_df.columns if col not in colunas_indice and col != "Total"]
                            if valores_pivot:
                                pivot_df['Total'] = pivot_df[valores_pivot].sum(axis=1)
                            else:
                                pivot_df['Total'] = 0
                            pivot_df.to_excel(writer, sheet_name="Pivot", index=False)

                        workbook = load_workbook(caminho_arquivo_material)
                        planilha = workbook["Pivot"]
                        max_row = planilha.max_row
                        max_col = planilha.max_column
                        if max_row > 1 and max_col > 1:
                            inicio_celula = planilha.cell(row=1, column=1).coordinate
                            fim_celula = planilha.cell(row=max_row, column=max_col).coordinate
                            tabela = Table(displayName=f"Table_Pivot", ref=f"{inicio_celula}:{fim_celula}")
                            estilo = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=True)
                            tabela.tableStyleInfo = estilo
                            planilha.add_table(tabela)
                            workbook.save(caminho_arquivo_material)
                            print(f"Tabela formatada no arquivo: {caminho_arquivo_material}")
                        else:
                            print(f"Planilha 'Pivot' vazia no arquivo {caminho_arquivo_material}. Formatação ignorada.")

                    except Exception as e:
                        print(f"Erro ao processar/salvar o arquivo para o material {material}: {e}")

            except Exception as e:
                print(f"Erro ao processar o arquivo {arquivo}: {e}")

        print("\nProcesso concluído!")

    except Exception as e:
        print(f"Erro geral durante a execução: {e}")


# Exemplo de uso (com parâmetros mais claros):
pasta_origem = r"\\srv-pt3\groups\02-Blades\02-Process Engineering\9. Projetos\4. Dashboard\Base de Dados\All - Por Ano"
pasta_destino = r"C:\Users\00082300\OneDrive - ENERCON\Área de Trabalho\BD_MES\Tabelas_Materiais"
coluna_filtro = "Work Center"
valores_desejados = ["E103_CUT","E103_FIN","E103_HKS","E103_KNF","E103_PF","E103_PP","E103_ROH","E103_SP","LEC-115","LEC-115N","PP_M001","PP_M002","PP_M003","PP_M004","PP_M005","PP_M006","PP_M007","PP_M008","PP_M009","PP_M010","PP_M011","PP_M012","PP_M013","PP_M014","PP_M015","PP_M017","PP_M018","PP_M019","PP_M020","QA0071","QA0071N","QA0200","QA0201","QA0202","QA0203","QA0204","QA0205","QA0206","QA0207","QA0208","QA0209","QA0210","QA0211","QA0212","QA0213","QA0214","QA0215","QA0216","QA0217","QA0218","QA0219","QA0220","QA0221","QA0222","QA0223","QA0224","QA0225","QA0226","QA0227","QA0228","QA0229","QA0230"]
colunas_indice = ["Material", "Material Description", "Serialnumber (PO Head)"]
colunas_valores = ["Total Confirmed Yield (MEINH)"]
coluna_operacao = "Operation Short Text"

gerar_tabelas_e_pivots(pasta_origem, pasta_destino, coluna_filtro, valores_desejados, colunas_indice, colunas_valores, coluna_operacao)