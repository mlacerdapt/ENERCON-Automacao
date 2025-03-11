import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo

# Caminho da pasta com os arquivos Excel
pasta = r"\\srv-pt3\groups\02-Blades\02-Process Engineering\9. Projetos\4. Dashboard\Base de Dados\All - Por Ano"
# Lista todos os arquivos dentro da pasta
arquivos = [os.path.join(pasta, arquivo) for arquivo in os.listdir(pasta)]

# Inicializa um DataFrame vazio para concatenar os dados
tabela_final = pd.DataFrame()

# Nome da coluna que será filtrada e a lista de valores desejados para o filtro
coluna_filtro = "Work Center"
valores_desejados = ["E103_CUT","E103_FIN","E103_HKS","E103_KNF","E103_PF","E103_PP","E103_ROH","E103_SP","LEC-115","LEC-115N","PP_M001","PP_M002","PP_M003","PP_M004","PP_M005","PP_M006","PP_M007","PP_M008","PP_M009","PP_M010","PP_M011","PP_M012","PP_M013","PP_M014","PP_M015","PP_M017","PP_M018","PP_M019","PP_M020","QA0071","QA0071N","QA0200","QA0201","QA0202","QA0203","QA0204","QA0205","QA0206","QA0207","QA0208","QA0209","QA0210","QA0211","QA0212","QA0213","QA0214","QA0215","QA0216","QA0217","QA0218","QA0219","QA0220","QA0221","QA0222","QA0223","QA0224","QA0225","QA0226","QA0227","QA0228","QA0229","QA0230"]

# Processar os arquivos e concatenar os dados filtrados
for arquivo in arquivos:
    try: # Tratamento de erro na leitura do arquivo
        df = pd.read_excel(arquivo, index_col=0)
        df_filtrado = df[df[coluna_filtro].isin(valores_desejados)]
        tabela_final = pd.concat([tabela_final, df_filtrado])
    except FileNotFoundError:
        print(f"Arquivo não encontrado: {arquivo}")
    except Exception as e:
        print(f"Erro ao processar o arquivo {arquivo}: {e}")


# Exporta a tabela final concatenada e filtrada para um arquivo Excel
output_file = r"C:\Users\00082300\OneDrive - ENERCON\Área de Trabalho\BD_MES\BD_103.xlsx"

try:
    with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
        if not tabela_final.empty:
            tabela_final.to_excel(writer, index=False, sheet_name="Consolidado")

            for material in tabela_final['Material'].unique():
                material_df = tabela_final[tabela_final['Material'] == material]
                if not material_df.empty:
                    material_df.to_excel(writer, sheet_name=str(material), index=True)

                    # Definir explicitamente as colunas de valores para a soma
                    values_columns = ["Total Confirmed Yield (MEINH)"] # Lista das colunas que serão somadas

                    pivot_df = pd.pivot_table(
                        tabela_final[tabela_final['Material'] == material],
                        values=values_columns, # Usando a lista de colunas aqui
                        index=["Material", "Material Description", "Serialnumber (PO Head)"],
                        columns="Operation Short Text",
                        aggfunc="sum"
                    ).reset_index()

                    if not pivot_df.empty:
                        # Achatando o MultiIndex das colunas
                        pivot_df.columns = [''.join(map(str, col)).strip() for col in pivot_df.columns.values]

                        # Adiciona a coluna de total somando as colunas de valores
                        values_columns = [col for col in pivot_df.columns if col not in ["Material", "Material Description", "Serialnumber (PO Head)", "Total"]] # Seleciona as colunas que não são index
                        if values_columns: # Verifica se a lista não está vazia
                            pivot_df['Total'] = pivot_df[values_columns].sum(axis=1)
                        else:
                            pivot_df['Total'] = 0 # Caso não tenha colunas para somar cria uma coluna total com valor 0

                        pivot_df.to_excel(writer, sheet_name=f"Pivot_{material}", index=False)

    print("Importação, filtro e concatenização concluídos com sucesso!")

    workbook = load_workbook(output_file)

    # Iterar SOMENTE sobre as planilhas "Pivot_*"
    for sheet_name in workbook.sheetnames:
        if sheet_name.startswith("Pivot_"):
            try:
                planilha = workbook[sheet_name]
                max_row = planilha.max_row
                max_col = planilha.max_column
                inicio_celula = planilha.cell(row=1, column=1).coordinate
                fim_celula = planilha.cell(row=max_row, column=max_col).coordinate
                tabela = Table(displayName=f"Table_{sheet_name}", ref=f"{inicio_celula}:{fim_celula}")
                estilo = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=True)
                tabela.tableStyleInfo = estilo
                planilha.add_table(tabela)
            except Exception as e:
                print(f"Erro ao formatar a planilha '{sheet_name}': {e}")

    workbook.save(output_file)
    print(f"Arquivo Excel '{output_file}' criado e formatado com sucesso!")

except Exception as e:
    print(f"Erro geral durante a execução do script: {e}")