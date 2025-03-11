import os 
import pandas as pd
from openpyxl import load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo


# Caminho da pasta com os arquivos Excel
pasta = r"\\srv-pt3\groups\02-Blades\02-Process Engineering\9. Projetos\4. Dashboard\Base de Dados\All - Por Ano"
# Lista todos os arquivos dentro da pasta
arquivos = [os.path.join(pasta, arquivo) for arquivo in os.listdir(pasta)]

# Inicializa um DataFrame vazio para concatenar os dados
tabela_final = pd.DataFrame()

# Nome da coluna que será filtrada e a lista de valores desejados para o filtro
coluna_filtro = "Work Center"
valores_desejados = ["PP_M005","QA0210","QA0232","PP_M020","PP_M018","QA0230","PP_M019","QA0231","PP_M017","QA0229","PP_M006","QA0211","PP_M013",
"QA0226","PP_M014","QA0227","PP_M015","QA0228","PP_M011","QA0221","PP_PA_N","QA0222","QA0223","PP_M004","QA0208","MM_PA_S","QA0209","PP_M008","QA0214","QA0215","PP_M012","QA0224","QA0225","PP_M007","QA0212","QA0213","PP_M010","QA0218","QA0219","QA0220","PP_M003","QA0206","QA0207","PP_M002","QA0203","QA0204","QA0205","PP_M009","QA0216","QA0217","PP_M001","QA0200","QA0201","QA0202"]  # Lista com os valores desejados

for arquivo in arquivos:
    df = pd.read_excel(arquivo, index_col=0)  # Lê o arquivo Excel
    df_filtrado = df[df[coluna_filtro].isin(valores_desejados)]  # Filtra os dados que possuem valores da lista
    tabela_final = pd.concat([tabela_final, df_filtrado])  # Concatena os dados filtrados na tabela final

# Exporta a tabela final concatenada e filtrada para um arquivo Excel

tabela_final.to_excel(r"\\srv-pt3\groups\02-Blades\04-Production\01 - Raw Parts\29 - Primary Parts\01 - Produção\09 - Farol MES\BD\BD_PP.xlsx")
print("Importação, filtro e concatenização concluídos com sucesso!")

df = pd.DataFrame(tabela_final)

# Exportar para Excel (criando o arquivo inicial)
arquivo_excel = r'\\srv-pt3\groups\02-Blades\04-Production\01 - Raw Parts\29 - Primary Parts\01 - Produção\09 - Farol MES\BD\BD_PP.xlsx'
df.to_excel(arquivo_excel, index=False, sheet_name='Planilha1')

# Abrir o arquivo com openpyxl para formatar
workbook = load_workbook(arquivo_excel)
planilha = workbook['Planilha1']

# Definir o intervalo dos dados como uma tabela
# Intervalo automático com base nos dados exportados
inicio_celula = planilha.cell(row=1, column=1).coordinate  # A1
fim_celula = planilha.cell(row=1 + len(df), column=len(df.columns)).coordinate  # Última célula

# Criar a tabela
tabela = Table(displayName="BD_MES", ref=f"{inicio_celula}:{fim_celula}")

# Definir o estilo da tabela
estilo = TableStyleInfo(
    name="TableStyleMedium9",  # Escolha o estilo desejado
    showFirstColumn=False,
    showLastColumn=False,
    showRowStripes=True,
    showColumnStripes=True
)
tabela.tableStyleInfo = estilo

# Adicionar a tabela à planilha
planilha.add_table(tabela)

# Salvar o arquivo Excel com a tabela formatada
workbook.save(arquivo_excel)
workbook.close()

print(f"Arquivo Excel '{arquivo_excel}' criado e formatado com sucesso!")