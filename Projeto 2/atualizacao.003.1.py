#-*- coding: UTF-8 -*-
import os
from time import sleep
import pyautogui
import pygetwindow as gw
import openpyxl
import schedule
import pyperclip
from datetime import datetime, timedelta
import shutil
import pandas as pd
import psutil
from openpyxl import load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from flask import Flask, jsonify, render_template
from threading import Timer
import win32com.client

blank = 'blank'
pa = 'all_PA'
pp = 'all_PP'
sa1 = 'stock_alerta'
sa2 = 'stock_alerta2'
sa3 = 'stock_alerta3'
mb52 = 'mb52'
me2n = 'me2n'
zmb52 ='zmb52'
backupr ="backup_resched"
backupcoois = "backupcoois"
backupzmb52 = "backupzmb52"
e115_bom = "E115_BOM"
e103_bom = "E103_BOM"
e175_bom = "E175_BOM"
p2p = "scm_p2p"
scm = "scm_verificaca"

def fechar_outlook():
    # Verifica todos os processos ativos
    for proc in psutil.process_iter(['pid', 'name']):
        # Se o nome do processo for 'OUTLOOK.EXE', fecha-o
        if proc.info['name'].lower() == 'outlook.exe':
            print(f"Fechando Outlook (PID: {proc.info['pid']})...")
            proc.terminate()  # Tenta fechar o processo de forma educada
            try:
                proc.wait(timeout=3)  # Espera por até 3 segundos para o processo fechar
                print("Outlook fechado com sucesso.")
            except psutil.TimeoutExpired:
                print("O Outlook não fechou dentro do tempo esperado. Forçando o fechamento.")
                proc.kill()  # Força o fechamento se não fechar dentro do tempo
            return
    print("Outlook não está aberto.")

def fechar_excel():
    try:
        # Conecta ao Excel se ele estiver aberto
        excel_app = win32com.client.Dispatch("Excel.Application")
        if excel_app.Workbooks.Count > 0:
            print("Salvando e fechando todas as planilhas abertas...")
            # Salva e fecha todas as planilhas abertas
            for workbook in excel_app.Workbooks:
                workbook.Save()
            excel_app.Quit()
            print("Excel fechado com sucesso após salvar as alterações.")
        else:
            print("Nenhuma planilha aberta. Fechando o Excel...")
            excel_app.Quit()
    except Exception as e:
        print(f"Erro ao tentar fechar o Excel: {e}")

    # Garantia extra de encerramento do processo
    for proc in psutil.process_iter(['pid', 'name']):
        if proc.info['name'].lower() == 'excel.exe':
            print(f"Forçando o fechamento do Excel (PID: {proc.info['pid']})...")
            proc.terminate()
            try:
                proc.wait(timeout=3)
                print("Excel encerrado com sucesso.")
            except psutil.TimeoutExpired:
                print("O Excel não respondeu a tempo. Forçando encerramento definitivo.")
                proc.kill()
            return

    print("Excel não está aberto.")

def check_message_on_screen(image_path, confidence=0.8):
    try:
        location = pyautogui.locateOnScreen(image_path, confidence=confidence)
        if location is not None:
            return True
        else:
            return False
    except pyautogui.ImageNotFoundException:
        return False
def verificar_msg():
    if __name__ == "__main__":
        image_path = r'C:\Users\00082300\Downloads\Projeto 2\export_conf1.png'  
        if check_message_on_screen(image_path):
            export_conf1 = pyautogui.locateCenterOnScreen('export_conf1.png', confidence=0.9)
            pyautogui.click(export_conf1[0],export_conf1[1], duration=1)
            print("A mensagem foi encontrada na tela.")
            sleep(10)
def enviar_erro(mensagem):
    # Função para enviar a mensagem de erro (implemente de acordo com sua necessidade)
    print(f"Erro: {mensagem}")

def verificar_imagem_na_tela(imagem_tela, max_tentativas=5, intervalo=5):
    tentativas = 0
    
    while tentativas < max_tentativas:
        try:
            # Tenta localizar a imagem na tela
            localizacao = pyautogui.locateCenterOnScreen(imagem_tela, confidence=0.9)
            
            # Se a imagem for encontrada, retorna a localização e sai do loop
            if localizacao:
                print(f"Imagem {imagem_tela} encontrada na tentativa {tentativas + 1}")
                return localizacao
        
        except pyautogui.ImageNotFoundException:
            pass

        # Incrementa as tentativas e espera o intervalo antes de tentar novamente
        tentativas += 1
        print(f"Tentativa {tentativas}, com a imagem {imagem_tela} falhou, aguardando {intervalo} segundos...")
        sleep(intervalo)

    # Se não encontrar a imagem após as tentativas, chama a função de erro
    enviar_erro(f"Imagem {imagem_tela} não encontrada após {max_tentativas} tentativas.")
    return None


def abrirsap():
    # ALL_PP
    #abrir o SAP
    os.startfile(r"C:\\Program Files (x86)\SAP\\FrontEnd\SAPgui\saplogon.exe")
    sleep(6)
def logarsap():
    #login do SAP
    imagem = 'botao_login_SAP.png'
    resultado = verificar_imagem_na_tela(imagem)
    if resultado:
        pyautogui.click(resultado[0],resultado[1], duration=1, clicks=2)
    else:
        print(f"Imagem {imagem} não encontrada.")
def abrircoois():
    #Abrir transação COOIS
    sleep(4)
    imagem = 'abrir_coois.png'
    resultado = verificar_imagem_na_tela(imagem)
    if resultado:
        pyautogui.click(resultado[0],resultado[1], duration=1, clicks=2)
    else:
        print(f"Imagem {imagem} não encontrada.")
    sleep(4)
def variantepp():
    #abrir variante(pp)
    pyautogui.hotkey('shift', 'f5')
    sleep(4)

    imagem = 'dig_variavel.png'
    resultado = verificar_imagem_na_tela(imagem)
    if resultado:
        pyautogui.click(resultado[0],resultado[1], duration=1, clicks=2)
    else:
        print(f"Imagem {imagem} não encontrada.")

    sleep(2)
    pyautogui.write(pp)
    
    imagem = 'conf_variavel.png'
    resultado = verificar_imagem_na_tela(imagem)
    if resultado:
        pyautogui.click(resultado[0],resultado[1], duration=1, clicks=2)
    else:
        print(f"Imagem {imagem} não encontrada.")
    sleep(1)
def abrirrelat():
    #Abrir relatório
    pyautogui.hotkey('f8')
    sleep(15)
def exportexcel():
    #abrir menu export
    imagem = 'abrir_menu_export.png'
    resultado = verificar_imagem_na_tela(imagem)
    if resultado:
        pyautogui.click(resultado[0],resultado[1], duration=1, clicks=1)
    else:
        print(f"Imagem {imagem} não encontrada.")

    sleep(1)

    #Exportar para excel
    imagem = 'export_excel.png'
    resultado = verificar_imagem_na_tela(imagem)
    if resultado:
        pyautogui.click(resultado[0],resultado[1], duration=1, clicks=1)
    else:
        print(f"Imagem {imagem} não encontrada.")
    sleep(1)

    imagem = 'export_excel_menu.png'
    resultado = verificar_imagem_na_tela(imagem)
    if resultado:
        pyautogui.click(resultado[0],resultado[1], duration=1, clicks=1)
    else:
        print(f"Imagem {imagem} não encontrada.")
    sleep(2)

    imagem = 'conf_export.png'
    resultado = verificar_imagem_na_tela(imagem)
    if resultado:
        pyautogui.click(resultado[0],resultado[1], duration=1, clicks=1)
    else:
        print(f"Imagem {imagem} não encontrada.")
    sleep(10)
def salvarexcelpp():
    #Salvar a planilha
    all_pp_salve = r'\\srv-pt3\groups\02-Blades\02-Process Engineering\9. Projetos\4. Dashboard\Base de Dados\all_pp.XLSX'
    pyautogui.write(all_pp_salve)
    pyautogui.hotkey('enter')
    sleep(4)
    pyautogui.hotkey('y')
    sleep(4)
    pyautogui.hotkey('alt','f4')
    sleep(2)
def sairtransacao():
    #sair transação
    pyautogui.hotkey('esc')
    sleep(3)
def variantepa():
    #abrir variante (PA)
    pyautogui.hotkey('shift', 'f5')
    sleep(4)
    imagem = 'dig_variavel.png'
    resultado = verificar_imagem_na_tela(imagem)
    if resultado:
        pyautogui.click(resultado[0],resultado[1], duration=1, clicks=2)
    else:
        print(f"Imagem {imagem} não encontrada.")
    sleep(2)
    pyautogui.write(pa)
    
    imagem = 'conf_variavel.png'
    resultado = verificar_imagem_na_tela(imagem)
    if resultado:
        pyautogui.click(resultado[0],resultado[1], duration=1, clicks=2)
    else:
        print(f"Imagem {imagem} não encontrada.")
    sleep(1)
def clicarno():
    #clicar no botão no para muitas linhas(superior a 5mil linhas)
    imagem = 'botao_no.png'
    resultado = verificar_imagem_na_tela(imagem)
    if resultado:
        pyautogui.click(resultado[0],resultado[1], duration=1, clicks=1)
    else:
        print(f"Imagem {imagem} não encontrada.")
    sleep(6)
def salvarexcelpa():
    #Salvar a planilha
    sleep(90)
    all_pa_salve = r'\\srv-pt3\groups\02-Blades\02-Process Engineering\9. Projetos\4. Dashboard\Base de Dados\All - Por Ano\All_2025.XLSX'
    pyautogui.write(all_pa_salve)
    sleep(10)
    pyautogui.hotkey('enter')
    sleep(10)
    pyautogui.hotkey('y')
    sleep(120)
    fechar_excel()
    sleep(10)
def abririq09():
    #Abrir transação IQ09 Stock
    imagem = 'abrir_iq09.png'
    resultado = verificar_imagem_na_tela(imagem)
    if resultado:
        pyautogui.click(resultado[0],resultado[1], duration=1, clicks=2)
    else:
        print(f"Imagem {imagem} não encontrada.")
    sleep(4)

    #abrir variante(pp)
    pyautogui.hotkey('shift', 'f5')
    sleep(4)
    imagem = 'dig_variavel.png'
    resultado = verificar_imagem_na_tela(imagem)
    if resultado:
        pyautogui.click(resultado[0],resultado[1], duration=1, clicks=2)
    else:
        print(f"Imagem {imagem} não encontrada.")
    sleep(4)
    pyautogui.write(pp)
    imagem = 'conf_variavel.png'
    resultado = verificar_imagem_na_tela(imagem)
    if resultado:
        pyautogui.click(resultado[0],resultado[1], duration=1, clicks=2)
    else:
        print(f"Imagem {imagem} não encontrada.")
    sleep(3)
    pyautogui.hotkey('f8')
    sleep(4)
def exportexceliq09():
    #Exportar excel
    imagem = 'export_iq09.png'
    resultado = verificar_imagem_na_tela(imagem)
    if resultado:
        pyautogui.click(resultado[0],resultado[1], duration=1, clicks=2)
    else:
        print(f"Imagem {imagem} não encontrada.")
    sleep(2)

    imagem = 'export_conf1.png'
    resultado = verificar_imagem_na_tela(imagem)
    if resultado:
        pyautogui.click(resultado[0],resultado[1], duration=1, clicks=1)
    else:
        print(f"Imagem {imagem} não encontrada.")
    sleep(3)
    pyautogui.hotkey('up')
    sleep(2)

    imagem = 'export_conf1.png'
    resultado = verificar_imagem_na_tela(imagem)
    if resultado:
        pyautogui.click(resultado[0],resultado[1], duration=1, clicks=1)
    else:
        print(f"Imagem {imagem} não encontrada.")
    sleep(2)

    imagem = 'export_conf1.png'
    resultado = verificar_imagem_na_tela(imagem)
    if resultado:
        pyautogui.click(resultado[0],resultado[1], duration=1, clicks=1)
    else:
        print(f"Imagem {imagem} não encontrada.")
    sleep(20)
    #Salvar excel
    imagem = 'salve_table.png'
    resultado = verificar_imagem_na_tela(imagem)
    if resultado:
        pyautogui.click(resultado[0],resultado[1], duration=1, clicks=1)
    else:
        print(f"Imagem {imagem} não encontrada.")
    sleep(2)
    pyautogui.hotkey('f12')
    sleep(3)
    all_pa_salve = r'\\srv-pt3\groups\02-Blades\02-Process Engineering\9. Projetos\4. Dashboard\Base de Dados\Stock_all_pp.xlsx'
    pyautogui.hotkey('backspace')
    sleep(3)
    pyautogui.hotkey('alt','n')
    sleep(3)
    pyautogui.write(all_pa_salve)
    pyautogui.hotkey('enter')
    sleep(1)
    pyautogui.hotkey('left')
    pyautogui.hotkey('enter')
    sleep(5)
    fechar_excel()
    sleep(1)
def fecharsap():
    #sair transação
    pyautogui.hotkey('alt','f4')
    sleep(1)
    imagem = 'yes_sair.png'
    resultado = verificar_imagem_na_tela(imagem)
    if resultado:
        pyautogui.click(resultado[0],resultado[1], duration=1, clicks=1)
    else:
        print(f"Imagem {imagem} não encontrada.")
    pyautogui.hotkey('enter')
    sleep(1)
    #fechar o SAP
    imagem = 'sair_x.png'
    resultado = verificar_imagem_na_tela(imagem)
    if resultado:
        pyautogui.click(resultado[0],resultado[1], duration=1, clicks=2)
    else:
        print(f"Imagem {imagem} não encontrada.")
def abrirtrans():
    #Abrir transação COOIS
    imagem = 'digitar_trans.png'
    resultado = verificar_imagem_na_tela(imagem)
    if resultado:
        pyautogui.click(resultado[0],resultado[1], duration=1, clicks=1)
    else:
        print(f"Imagem {imagem} não encontrada.")

    sleep(1)
def abrirmb52():
    pyautogui.write(mb52)
    pyautogui.hotkey('enter')
    sleep(2)
    #abrir variante(sa1)
    pyautogui.hotkey('shift', 'f5')
    sleep(4)
    imagem = 'dig_variavel.png'
    resultado = verificar_imagem_na_tela(imagem)
    if resultado:
        pyautogui.click(resultado[0],resultado[1], duration=1, clicks=1)
    else:
        print(f"Imagem {imagem} não encontrada.")
    sleep(4)
    pyautogui.write(sa1)
    imagem = 'conf_variavel.png'
    resultado = verificar_imagem_na_tela(imagem)
    if resultado:
        pyautogui.click(resultado[0],resultado[1], duration=1, clicks=2)
    else:
        print(f"Imagem {imagem} não encontrada.")
    sleep(3)
    pyautogui.hotkey('f8')
    sleep(4)
def abrirme2n():
    pyautogui.write(me2n)
    pyautogui.hotkey('enter')
    sleep(2)
    #abrir variante(sa1)
    pyautogui.hotkey('shift', 'f5')
    sleep(4)
    imagem = 'dig_variavel.png'
    resultado = verificar_imagem_na_tela(imagem)
    if resultado:
        pyautogui.click(resultado[0],resultado[1], duration=1, clicks=1)
    else:
        print(f"Imagem {imagem} não encontrada.")
    sleep(4)
    pyautogui.write(sa2)

    imagem = 'conf_variavel.png'
    resultado = verificar_imagem_na_tela(imagem)
    if resultado:
        pyautogui.click(resultado[0],resultado[1], duration=1, clicks=1)
    else:
        print(f"Imagem {imagem} não encontrada.")

    sleep(3)

    imagem = 'acesso_variante.png'
    resultado = verificar_imagem_na_tela(imagem)
    if resultado:
        pyautogui.click(resultado[0],resultado[1], duration=1, clicks=1)
    else:
        print(f"Imagem {imagem} não encontrada.")
    sleep(3)

    imagem = 'Conf_alteracao.png'
    resultado = verificar_imagem_na_tela(imagem)
    if resultado:
        pyautogui.click(resultado[0],resultado[1], duration=1, clicks=1)
    else:
        print(f"Imagem {imagem} não encontrada.")
    sleep(3)
    pyautogui.hotkey('f8')
    sleep(30)
def abrirzmb52():
    pyautogui.write(zmb52)
    pyautogui.hotkey('enter')
    sleep(2)
    #abrir variante(sa1)
    pyautogui.hotkey('shift', 'f5')
    sleep(4)

    imagem = 'dig_variavel.png'
    resultado = verificar_imagem_na_tela(imagem)
    if resultado:
        pyautogui.click(resultado[0],resultado[1], duration=1, clicks=1)
    else:
        print(f"Imagem {imagem} não encontrada.")
    sleep(4)
    pyautogui.write(sa3)

    imagem = 'conf_variavel.png'
    resultado = verificar_imagem_na_tela(imagem)
    if resultado:
        pyautogui.click(resultado[0],resultado[1], duration=1, clicks=2)
    else:
        print(f"Imagem {imagem} não encontrada.")
    sleep(3)
    pyautogui.hotkey('f8')
    sleep(60)
def trans_blank():
    #abrir transação
    imagem = 'mb51_serial.png'
    resultado = verificar_imagem_na_tela(imagem)
    if resultado:
        pyautogui.click(resultado[0],resultado[1], duration=1, clicks=2)
    else:
        print(f"Imagem {imagem} não encontrada.")
    sleep(2)

    #abrir variante(blank)
    pyautogui.hotkey('shift', 'f5')
    sleep(4)

    imagem = 'janelaexportar.png'
    resultado = verificar_imagem_na_tela(imagem)
    if resultado:
        pyautogui.click(resultado[0],resultado[1], duration=1, clicks=2)
    else:
        print(f"Imagem {imagem} não encontrada.")
    sleep(2)
    pyautogui.write(blank)
    imagem = 'conf_variavel.png'
    resultado = verificar_imagem_na_tela(imagem)
    if resultado:
        pyautogui.click(resultado[0],resultado[1], duration=1, clicks=2)
    else:
        print(f"Imagem {imagem} não encontrada.")
    sleep(3)
    pyautogui.hotkey('f8')
    sleep(10)
def export_menu():
    imagem = 'menu-list.png'
    resultado = verificar_imagem_na_tela(imagem)
    if resultado:
        pyautogui.click(resultado[0],resultado[1], duration=1, clicks=1)
    else:
        print(f"Imagem {imagem} não encontrada.")
    sleep(1)

    menu_export2= pyautogui.locateCenterOnScreen('menu-export.png', confidence=0.9)
    pyautogui.click(menu_export2[0],menu_export2[1], duration=1)
    sleep(1)
    menusheet= pyautogui.locateCenterOnScreen('menu_sheet.png', confidence=0.9)
    pyautogui.click(menusheet[0],menusheet[1], duration=1)
    sleep(1)
    confexport = pyautogui.locateCenterOnScreen('conf_export.png', confidence=0.9)
    pyautogui.click(confexport[0],confexport[1], duration=1)
    sleep(30)
def export_blank():
    #Salvar a planilha
    all_blank = r'\\srv-pt3\groups\02-Blades\02-Process Engineering\9. Projetos\4. Dashboard\Base de Dados\Serialnumber_blank.xlsx'
    pyautogui.write(all_blank)
    pyautogui.hotkey('enter')
    sleep(10)
    pyautogui.hotkey('y')
    sleep(20)
    fechar_excel()
    sleep(2)
def export_mb52SA():
    #Salvar a planilha
    mb52SA = r'\\srv-pt3\groups\02-Blades\05-Warehouse\08. Análise semanal de níveis de stock\Base de Dados\SAP\MB52_All.XLSX'
    pyperclip.copy(mb52SA)
    pyautogui.hotkey('ctrl', 'v')
    pyautogui.hotkey('enter')
    sleep(4)
    pyautogui.hotkey('y')
    sleep(60)
    fechar_excel()
    sleep(2)
def export_me2nSA():
    imagem = 'abrirdatas.png'
    resultado = verificar_imagem_na_tela(imagem)
    if resultado:
        pyautogui.click(resultado[0],resultado[1], duration=1, clicks=2)
    else:
        print(f"Imagem {imagem} não encontrada.")
    sleep(6)

    imagem = 'devildate.png'
    resultado = verificar_imagem_na_tela(imagem)
    if resultado:
        pyautogui.click(resultado[0],resultado[1], duration=1, clicks=1, button='right')
    else:
        print(f"Imagem {imagem} não encontrada.")
    sleep(1)

    imagem = 'acendente.png'
    resultado = verificar_imagem_na_tela(imagem)
    if resultado:
        pyautogui.click(resultado[0],resultado[1], duration=1, clicks=1)
    else:
        print(f"Imagem {imagem} não encontrada.")
    sleep(1)

    imagem = 'export_iq09.png'
    resultado = verificar_imagem_na_tela(imagem)
    if resultado:
        pyautogui.click(resultado[0],resultado[1], duration=1, clicks=1)
    else:
        print(f"Imagem {imagem} não encontrada.")
    sleep(2)

    imagem = 'conf_export.png'
    resultado = verificar_imagem_na_tela(imagem)
    if resultado:
        pyautogui.click(resultado[0],resultado[1], duration=1, clicks=1)
    else:
        print(f"Imagem {imagem} não encontrada.")
    sleep(6)
    #Salvar a planilha
    me2nSA = r'\\srv-pt3\groups\02-Blades\05-Warehouse\08. Análise semanal de níveis de stock\Base de Dados\SAP\ME2N_All.XLSX'
    pyperclip.copy(me2nSA)
    pyautogui.hotkey('ctrl', 'v')
    pyautogui.hotkey('enter')
    sleep(4)
    pyautogui.hotkey('y')
    sleep(60)
    fechar_excel()
    sleep(2)
def export_zmb52SA():
    pyautogui.hotkey('ctrl','shift','f7')
    sleep(2)

    imagem = 'conf_export.png'
    resultado = verificar_imagem_na_tela(imagem)
    if resultado:
        pyautogui.click(resultado[0],resultado[1], duration=1, clicks=1)
    else:
        print(f"Imagem {imagem} não encontrada.")
    sleep(6)

    #Salvar a planilha
    zmb52SA = r'\\srv-pt3\groups\02-Blades\05-Warehouse\08. Análise semanal de níveis de stock\Base de Dados\SAP\ZMB52_All.XLSX'
    pyperclip.copy(zmb52SA)
    pyautogui.hotkey('ctrl', 'v')
    pyautogui.hotkey('enter')
    sleep(4)
    pyautogui.hotkey('y')
    sleep(60)
    fechar_excel()
    sleep(2)
def controle_corte():
    os.startfile(r"\\srv-pt3\groups\02-Blades\02-Process Engineering\9. Projetos\4. Dashboard\Base de Dados\CONTROLE CORTE.xlsm")
    sleep(45)
    fechar_excel()
    print('Controle de Corte atualizado!')
    sleep(10)
def atualização_BD():
    
    os.startfile(r'\\srv-pt3\groups\02-Blades\02-Process Engineering\9. Projetos\4. Dashboard\Base de Dados\BD_2024_dashboard.xlsm')
    sleep(30)
    pyautogui.hotkey('ctrl', 'alt','f5')
    pyautogui.hotkey('left')
    pyautogui.hotkey('enter')
    print('Atualizando powerquery')
    sleep(120)
    print('Finalizado o prazo de atualização PowerQuery')
    pyautogui.hotkey('ctrl', 'q')
    print('Atualizando Macro')
    sleep(250)
    print('Finalziado a atualização da Macro')
def abrirjanelas():
    #abrir 3 janelas
    sleep(5)
    pyautogui.hotkey('ctrl', 'n')
    sleep(5)
    pyautogui.hotkey('ctrl', 'n')
    sleep(5)
def focus_window(title):
    windows = gw.getWindowsWithTitle(title)
    if windows:
        window = windows[0]
        window.activate()
    else:
        print(f'Janela com título "{title}" não encontrada.')
def atualização_SA():
    os.startfile(r'\\srv-pt3\groups\02-Blades\05-Warehouse\08. Análise semanal de níveis de stock\Base de Dados\Teste\2. PTA3_stock alerta(MACRO).xlsm')
    sleep(15)
    pyautogui.hotkey('ctrl', 'w')
    print('Atualizando Macro')
    sleep(120)
    pyautogui.hotkey('enter')
    print('Finalziado a atualização da Macro PTA3')
    pyautogui.hotkey('alt', 'f4')
    sleep(4)
    pyautogui.hotkey('g')
    print('Finalziado arquivo PTA3')
    sleep(4)
    os.startfile(r'\\srv-pt3\groups\02-Blades\05-Warehouse\08. Análise semanal de níveis de stock\Base de Dados\Teste\4. STOCK ALERTA PTA0 _PTA3.xlsm')
    sleep(15)
    pyautogui.hotkey('ctrl', 'w')
    print('Atualizando Macro')
    sleep(120)
    pyautogui.hotkey('enter')
    print('Finalziado a atualização da Macro PTA0')
    pyautogui.hotkey('alt', 'f4')
    sleep(4)
    pyautogui.hotkey('g')
    print('Finalziado arquivo PTA0')

def atualizarfarol_PP():
    # Obter a data e hora atuais
    data_hora_atual = datetime.now()
    # Formatar a data e hora em uma string
    nomearquivo = data_hora_atual.strftime("%d-%m-%y %H%M%S")
    # Exibir a data e hora formatadas
    print("Data e hora atuais formatadas:", nomearquivo)
    # Caminho da pasta com os arquivos Excel
    pasta = r"\\srv-pt3\groups\02-Blades\02-Process Engineering\9. Projetos\4. Dashboard\Base de Dados\All - Por Ano"
    # Lista todos os arquivos dentro da pasta
    arquivos = [os.path.join(pasta, arquivo) for arquivo in os.listdir(pasta)]

    # Inicializa um DataFrame vazio para concatenar os dados
    tabela_final = pd.DataFrame()

    # Nome da coluna que será filtrada e a lista de valores desejados para o filtro
    coluna_filtro = "Work Center"
    valores_desejados = ["PP_M005","QA0210","QA0232","PP_M020","PP_M018","QA0230","PP_M019","QA0231","PP_M017","QA0229","PP_M006","QA0211","PP_M013",
    "QA0226","PP_M014","QA0227","PP_M015","QA0228","PP_M011","QA0221","PP_PA_N","QA0222","QA0223","PP_M004","QA0208","MM_PA_S","QA0209","PP_M008","QA0214","QA0215","PP_M012","QA0224","QA0225","PP_M007","QA0212","QA0213","PP_M010","QA0218","QA0219","QA0220","PP_M003","QA0206","QA0207","PP_M002","QA0203","QA0204","QA0205","PP_M009","QA0216","QA0217","PP_M001","QA0200","QA0201","QA0202"]  # Lista com os valores desejados

    for arquivo in arquivos:
        df = pd.read_excel(arquivo, index_col=0)  # Lê o arquivo Excel
        df_filtrado = df[df[coluna_filtro].isin(valores_desejados)]  # Filtra os dados que possuem valores da lista
        tabela_final = pd.concat([tabela_final, df_filtrado])  # Concatena os dados filtrados na tabela final

    # Exporta a tabela final concatenada e filtrada para um arquivo Excel

    tabela_final.to_excel(r"\\srv-pt3\groups\02-Blades\04-Production\01 - Raw Parts\29 - Primary Parts\01 - Produção\09 - Farol MES\BD\BD_PP.xlsx")
    print("Importação, filtro e concatenização concluídos com sucesso!")

    df = pd.DataFrame(tabela_final)

    # Exportar para Excel (criando o arquivo inicial)
    arquivo_excel = r'\\srv-pt3\groups\02-Blades\04-Production\01 - Raw Parts\29 - Primary Parts\01 - Produção\09 - Farol MES\BD\BD_PP.xlsx'
    df.to_excel(arquivo_excel, index=False, sheet_name='Planilha1')

    # Abrir o arquivo com openpyxl para formatar
    workbook = load_workbook(arquivo_excel)
    planilha = workbook['Planilha1']

    # Definir o intervalo dos dados como uma tabela
    # Intervalo automático com base nos dados exportados
    inicio_celula = planilha.cell(row=1, column=1).coordinate  # A1
    fim_celula = planilha.cell(row=1 + len(df), column=len(df.columns)).coordinate  # Última célula

    # Criar a tabela
    tabela = Table(displayName="Tabela1", ref=f"{inicio_celula}:{fim_celula}")

    # Definir o estilo da tabela
    estilo = TableStyleInfo(
        name="TableStyleMedium9",  # Escolha o estilo desejado
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=True
    )
    tabela.tableStyleInfo = estilo

    # Adicionar a tabela à planilha
    planilha.add_table(tabela)

    # Salvar o arquivo Excel com a tabela formatada
    workbook.save(arquivo_excel)
    workbook.close()

    print(f"Arquivo Excel '{arquivo_excel}' criado e formatado com sucesso!")
    tempo = datetime.now() - data_hora_atual
    print("Finalizando em:", tempo)

def atualizar_bd_PP():
        # Obter a data e hora atuais
    data_hora_atual = datetime.now()
    # Formatar a data e hora em uma string
    nomearquivo = data_hora_atual.strftime("%d-%m-%y %H%M%S")
    # Exibir a data e hora formatadas
    print("Data e hora atuais formatadas:", nomearquivo)
    fechar_outlook()
    abrirsap()
    sleep(2)
    verificar_msg()
    sleep(2)
    logarsap()
    sleep(2)
    verificar_msg()
    sleep(2)
    abrircoois()
    variantepa()
    abrirrelat()
    clicarno()
    print('Iniciado COOIS PA')
    sleep(350)
    controle_corte()
    sleep(20)
    exportexcel()
    salvarexcelpa()
    sairtransacao()
    sairtransacao()
    print('Arquivo gerado! Transação: COOIS')
    sleep(2)
    fecharsap()
    atualizarfarol_PP()
    tempo = datetime.now() - data_hora_atual
    print("Finalizando em:", tempo)
    print('**************BD Relatório Status de Produção atualizado com sucesso!*************')

def abrirrescheduling():
    # Abrir o aplicativo
    os.startfile(r"C:\Users\00082300\Downloads\Projeto 2\atalhos\mm_utr.SAP")
    sleep(10)

    # Tentar encontrar a janela aberta
    windows = gw.getWindowsWithTitle('ENP')

    if windows:
        app_window = windows[0]
        app_window.maximize()
    else:
        print("Janela do aplicativo não encontrada")
def abrirbackupcoois():
    # Abrir o aplicativo
    os.startfile(r"C:\Users\00082300\Downloads\Projeto 2\atalhos\coois.SAP")
    sleep(10)

    # Tentar encontrar a janela aberta
    windows = gw.getWindowsWithTitle('ENP')

    if windows:
        app_window = windows[0]
        app_window.maximize()
    else:
        print("Janela do aplicativo não encontrada")   
def abrirbackupzmb52():
    # Abrir o aplicativo
    os.startfile(r"C:\Users\00082300\Downloads\Projeto 2\atalhos\zmb52.SAP")
    sleep(10)

    # Tentar encontrar a janela aberta
    windows = gw.getWindowsWithTitle('ENP')

    if windows:
        app_window = windows[0]
        app_window.maximize()
    else:
        print("Janela do aplicativo não encontrada")  
def abrirbackupzbom():
    # Abrir o aplicativo
    os.startfile(r"C:\Users\00082300\Downloads\Projeto 2\atalhos\zbom.SAP")
    sleep(10)

    # Tentar encontrar a janela aberta
    windows = gw.getWindowsWithTitle('ENP')

    if windows:
        app_window = windows[0]
        app_window.maximize()
    else:
        print("Janela do aplicativo não encontrada")  
def abrirbackupp2p():
    # Abrir o aplicativo
    os.startfile(r"C:\Users\00082300\Downloads\Projeto 2\atalhos\p2p.SAP")
    sleep(10)

    # Tentar encontrar a janela aberta
    windows = gw.getWindowsWithTitle('ENP')

    if windows:
        app_window = windows[0]
        app_window.maximize()
    else:
        print("Janela do aplicativo não encontrada")  

def exerescheduling():
    pyautogui.hotkey('shift', 'f5')
    sleep(4)

    imagem = 'janelaexportar.png'
    resultado = verificar_imagem_na_tela(imagem)
    if resultado:
        pyautogui.click(resultado[0],resultado[1], duration=1, clicks=2)
    else:
        print(f"Imagem {imagem} não encontrada.")
    sleep(2)
    pyautogui.write(backupr)

    imagem = 'conf_variavel.png'
    resultado = verificar_imagem_na_tela(imagem)
    if resultado:
        pyautogui.click(resultado[0],resultado[1], duration=1, clicks=2)
    else:
        print(f"Imagem {imagem} não encontrada.")
    sleep(3)
    pyautogui.hotkey('f8')
    sleep(80)
def execoois():
    # Obter a data e hora atuais
    data_hora_atual = datetime.now()
    # Formatar a data e hora em uma string
    nomearquivo = data_hora_atual.strftime("%d-%m-%y %H%M%S")

    # Obter a data de inicio e fim do periodo de pesquisa
    data_inicio = datetime.now() - timedelta(days=35)
    data_fim = datetime.now() + timedelta(days=35)
    # Formatar a data e hora em uma string
    datainicio = data_inicio.strftime("%d%m%Y")
    datafim = data_fim.strftime("%d%m%Y")
    pyautogui.hotkey('shift', 'f5')
    sleep(4)

    imagem = 'janelaexportar.png'
    resultado = verificar_imagem_na_tela(imagem)
    if resultado:
        pyautogui.click(resultado[0],resultado[1], duration=1, clicks=2)
    else:
        print(f"Imagem {imagem} não encontrada.")

    sleep(4)
    pyautogui.write(backupcoois)

    imagem = 'conf_variavel.png'
    resultado = verificar_imagem_na_tela(imagem)
    if resultado:
        pyautogui.click(resultado[0],resultado[1], duration=1, clicks=2)
    else:
        print(f"Imagem {imagem} não encontrada.")
    sleep(4)

    imagem = 'datestart.png'
    resultado = verificar_imagem_na_tela(imagem)
    if resultado:
        pyautogui.click(resultado[0],resultado[1], duration=1, clicks=2)
    else:
        print(f"Imagem {imagem} não encontrada.")
    sleep(4)

    imagem = 'X_vemerlho.png'
    resultado = verificar_imagem_na_tela(imagem)
    if resultado:
        pyautogui.click(resultado[0],resultado[1], duration=1, clicks=1)
    else:
        print(f"Imagem {imagem} não encontrada.")
    sleep(6)
    pyautogui.write(datainicio)

    imagem = 'datefinish.png'
    resultado = verificar_imagem_na_tela(imagem)
    if resultado:
        pyautogui.click(resultado[0],resultado[1], duration=1, clicks=2)
    else:
        print(f"Imagem {imagem} não encontrada.")
    sleep(4)

    imagem = 'X_vemerlho.png'
    resultado = verificar_imagem_na_tela(imagem)
    if resultado:
        pyautogui.click(resultado[0],resultado[1], duration=1, clicks=1)
    else:
        print(f"Imagem {imagem} não encontrada.")
    sleep(4)
    pyautogui.write(datafim)

    pyautogui.hotkey('f8')
    sleep(80)
def exezmb52():
    pyautogui.hotkey('shift', 'f5')
    sleep(4)

    imagem = 'janelaexportar.png'
    resultado = verificar_imagem_na_tela(imagem)
    if resultado:
        pyautogui.click(resultado[0],resultado[1], duration=1, clicks=2)
    else:
        print(f"Imagem {imagem} não encontrada.")
    sleep(4)
    pyautogui.write(backupzmb52)

    imagem = 'conf_variavel.png'
    resultado = verificar_imagem_na_tela(imagem)
    if resultado:
        pyautogui.click(resultado[0],resultado[1], duration=1, clicks=2)
    else:
        print(f"Imagem {imagem} não encontrada.")
    sleep(4)
    pyautogui.hotkey('f8')
    sleep(80)
def exep2p():
    pyautogui.hotkey('shift', 'f5')
    sleep(4)

    imagem = 'janelaexportar.png'
    resultado = verificar_imagem_na_tela(imagem)
    if resultado:
        pyautogui.click(resultado[0],resultado[1], duration=1, clicks=2)
    else:
        print(f"Imagem {imagem} não encontrada.")

    sleep(4)
    pyautogui.write(p2p)

    imagem = 'conf_variavel.png'
    resultado = verificar_imagem_na_tela(imagem)
    if resultado:
        pyautogui.click(resultado[0],resultado[1], duration=1, clicks=2)
    else:
        print(f"Imagem {imagem} não encontrada.")
    sleep(6)
    pyautogui.hotkey('f8')
    sleep(80)

def exezbom_e115():
    pyautogui.hotkey('shift', 'f5')
    sleep(4)
    imagem = 'janelaexportar.png'
    resultado = verificar_imagem_na_tela(imagem)
    if resultado:
        pyautogui.click(resultado[0],resultado[1], duration=1, clicks=2)
    else:
        print(f"Imagem {imagem} não encontrada.")
    sleep(4)
    pyautogui.write(e115_bom)
    imagem = 'conf_variavel.png'
    resultado = verificar_imagem_na_tela(imagem)
    if resultado:
        pyautogui.click(resultado[0],resultado[1], duration=1, clicks=2)
    else:
        print(f"Imagem {imagem} não encontrada.")
    sleep(6)
    pyautogui.hotkey('f8')
    sleep(80)
def exezbom_e175():
    pyautogui.hotkey('shift', 'f5')
    sleep(4)
    imagem = 'janelaexportar.png'
    resultado = verificar_imagem_na_tela(imagem)
    if resultado:
        pyautogui.click(resultado[0],resultado[1], duration=1, clicks=2)
    else:
        print(f"Imagem {imagem} não encontrada.")
    sleep(4)
    pyautogui.write(e175_bom)

    imagem = 'conf_variavel.png'
    resultado = verificar_imagem_na_tela(imagem)
    if resultado:
        pyautogui.click(resultado[0],resultado[1], duration=1, clicks=2)
    else:
        print(f"Imagem {imagem} não encontrada.")
    sleep(6)
    pyautogui.hotkey('f8')
    sleep(80)
def exezbom_e103():
    pyautogui.hotkey('shift', 'f5')
    sleep(4)

    imagem = 'janelaexportar.png'
    resultado = verificar_imagem_na_tela(imagem)
    if resultado:
        pyautogui.click(resultado[0],resultado[1], duration=1, clicks=2)
    else:
        print(f"Imagem {imagem} não encontrada.")
    sleep(4)
    pyautogui.write(e103_bom)

    imagem = 'conf_variavel.png'
    resultado = verificar_imagem_na_tela(imagem)
    if resultado:
        pyautogui.click(resultado[0],resultado[1], duration=1, clicks=2)
    else:
        print(f"Imagem {imagem} não encontrada.")
    sleep(4)
    pyautogui.hotkey('f8')
    sleep(80)

def export_backupr():
    pyautogui.hotkey('shift','f9')
    sleep(4)
    imagem = 'conf_export.png'
    resultado = verificar_imagem_na_tela(imagem)
    if resultado:
        pyautogui.click(resultado[0],resultado[1], duration=1, clicks=1)
    else:
        print(f"Imagem {imagem} não encontrada.")
    sleep(60)
    arquivo = r'O:\02-Blades\02-Process Engineering\9. Projetos\22. Backup diario\IT-Breakdown Open Purchase Orders Actual Status'
    pyautogui.write(arquivo)
    pyautogui.hotkey('enter')
    sleep(6)
    pyautogui.hotkey('y')
    sleep(80)
    fechar_excel()
    sleep(10)

    imagem = 'sair_x.png'
    resultado = verificar_imagem_na_tela(imagem)
    if resultado:
        pyautogui.click(resultado[0],resultado[1], duration=1, clicks=1)
    else:
        print(f"Imagem {imagem} não encontrada.")
    sleep(4)

    imagem = 'yes_sair.png'
    resultado = verificar_imagem_na_tela(imagem)
    if resultado:
        pyautogui.click(resultado[0],resultado[1], duration=1, clicks=1)
    else:
        print(f"Imagem {imagem} não encontrada.")
def export_backupcoois():
    #abrir menu export
    imagem = 'abrir_menu_export.png'
    resultado = verificar_imagem_na_tela(imagem)
    if resultado:
        pyautogui.click(resultado[0],resultado[1], duration=1, clicks=1)
    else:
        print(f"Imagem {imagem} não encontrada.")
    sleep(4)

    #Exportar para excel
    imagem = 'export_excel.png'
    resultado = verificar_imagem_na_tela(imagem)
    if resultado:
        pyautogui.click(resultado[0],resultado[1], duration=1, clicks=1)
    else:
        print(f"Imagem {imagem} não encontrada.")
    sleep(4)

    imagem = 'export_excel_menu.png'
    resultado = verificar_imagem_na_tela(imagem)
    if resultado:
        pyautogui.click(resultado[0],resultado[1], duration=1, clicks=1)
    else:
        print(f"Imagem {imagem} não encontrada.")
    sleep(4)

    imagem = 'conf_export.png'
    resultado = verificar_imagem_na_tela(imagem)
    if resultado:
        pyautogui.click(resultado[0],resultado[1], duration=1, clicks=1)
    else:
        print(f"Imagem {imagem} não encontrada.")
    sleep(60)
    arquivo = r'O:\02-Blades\02-Process Engineering\9. Projetos\22. Backup diario\IT-Breakdown Production Orders_Converted and Planned'
    pyautogui.write(arquivo)
    pyautogui.hotkey('enter')
    sleep(6)
    pyautogui.hotkey('y')
    sleep(80)
    fechar_excel()
    sleep(10)
    imagem = 'sair_x.png'
    resultado = verificar_imagem_na_tela(imagem)
    if resultado:
        pyautogui.click(resultado[0],resultado[1], duration=1, clicks=2)
    else:
        print(f"Imagem {imagem} não encontrada.")
    sleep(4)

    imagem = 'yes_sair.png'
    resultado = verificar_imagem_na_tela(imagem)
    if resultado:
        pyautogui.click(resultado[0],resultado[1], duration=1, clicks=2)
    else:
        print(f"Imagem {imagem} não encontrada.")

def export_backupzmb52():
    pyautogui.hotkey('ctrl','shift','f7')
    sleep(4)
    imagem = 'conf_export.png'
    resultado = verificar_imagem_na_tela(imagem)
    if resultado:
        pyautogui.click(resultado[0],resultado[1], duration=1, clicks=1)
    else:
        print(f"Imagem {imagem} não encontrada.")
    sleep(60)
    arquivo = r'O:\02-Blades\02-Process Engineering\9. Projetos\22. Backup diario\IT-Breakdown Current Stocks PTA0_PTA3 '
    pyautogui.write(arquivo)
    pyautogui.hotkey('enter')
    sleep(6)
    pyautogui.hotkey('y')
    sleep(80)
    fechar_excel()
    sleep(10)

    imagem = 'sair_x.png'
    resultado = verificar_imagem_na_tela(imagem)
    if resultado:
        pyautogui.click(resultado[0],resultado[1], duration=1, clicks=1)
    else:
        print(f"Imagem {imagem} não encontrada.")
    sleep(4)

    imagem = 'yes_sair.png'
    resultado = verificar_imagem_na_tela(imagem)
    if resultado:
        pyautogui.click(resultado[0],resultado[1], duration=1, clicks=1)
    else:
        print(f"Imagem {imagem} não encontrada.")

def export_backup2p():
    #Exportar para excel
    imagem = 'export_excel.png'
    resultado = verificar_imagem_na_tela(imagem)
    if resultado:
        pyautogui.click(resultado[0],resultado[1], duration=1, clicks=1)
    else:
        print(f"Imagem {imagem} não encontrada.")
    sleep(6)

    imagem = 'export_excel_menu.png'
    resultado = verificar_imagem_na_tela(imagem)
    if resultado:
        pyautogui.click(resultado[0],resultado[1], duration=1, clicks=1)
    else:
        print(f"Imagem {imagem} não encontrada.")
    sleep(6)

    imagem = 'conf_export.png'
    resultado = verificar_imagem_na_tela(imagem)
    if resultado:
        pyautogui.click(resultado[0],resultado[1], duration=1, clicks=1)
    else:
        print(f"Imagem {imagem} não encontrada.")
    sleep(60)
    arquivo = r'O:\02-Blades\02-Process Engineering\9. Projetos\22. Backup diario\IT-Breakdown Global Material Responsibles'
    pyautogui.write(arquivo)
    pyautogui.hotkey('enter')
    sleep(6)
    pyautogui.hotkey('y')
    sleep(80)
    fechar_excel()
    sleep(10)
    imagem = 'sair_x.png'
    resultado = verificar_imagem_na_tela(imagem)
    if resultado:
        pyautogui.click(resultado[0],resultado[1], duration=1, clicks=1)
    else:
        print(f"Imagem {imagem} não encontrada.")
    sleep(4)

    imagem = 'yes_sair.png'
    resultado = verificar_imagem_na_tela(imagem)
    if resultado:
        pyautogui.click(resultado[0],resultado[1], duration=1, clicks=1)
    else:
        print(f"Imagem {imagem} não encontrada.")
def export_backupzbom_e115():
    arquivo = r'O:\02-Blades\02-Process Engineering\9. Projetos\22. Backup diario\IT-Breakdown BOM_Standard Grey Blade E115'
    pyautogui.write(arquivo)
    pyautogui.hotkey('enter')
    sleep(4)
    pyautogui.hotkey('y')
    sleep(80)
    fechar_excel()
    sleep(10)
    pyautogui.hotkey('esc')
    sleep(3)
def export_backupzbom_e175():
    arquivo = r'O:\02-Blades\02-Process Engineering\9. Projetos\22. Backup diario\IT-Breakdown BOM_Standard Grey Blade E175'
    pyautogui.write(arquivo)
    pyautogui.hotkey('enter')
    sleep(4)
    pyautogui.hotkey('y')
    sleep(80)
    fechar_excel()
    sleep(10)
    pyautogui.hotkey('esc')
    sleep(4)
def export_backupzbom_e103():
    arquivo = r'O:\02-Blades\02-Process Engineering\9. Projetos\22. Backup diario\IT-Breakdown BOM_Standard Grey Blade E103'
    pyautogui.write(arquivo)
    pyautogui.hotkey('enter')
    sleep(6)
    pyautogui.hotkey('y')
    sleep(80)
    fechar_excel()
    sleep(10)
    imagem = 'sair_x.png'
    resultado = verificar_imagem_na_tela(imagem)
    if resultado:
        pyautogui.click(resultado[0],resultado[1], duration=1, clicks=1)
    else:
        print(f"Imagem {imagem} não encontrada.")
    sleep(4)

    imagem = 'yes_sair.png'
    resultado = verificar_imagem_na_tela(imagem)
    if resultado:
        pyautogui.click(resultado[0],resultado[1], duration=1, clicks=1)
    else:
        print(f"Imagem {imagem} não encontrada.")
def copiar_arquivos(pasta_origem, pasta_destino):
    try:
        if not os.path.exists(pasta_destino):
            os.makedirs(pasta_destino)

        arquivos = os.listdir(pasta_origem)
        for arquivo in arquivos:
            caminho_origem = os.path.join(pasta_origem, arquivo)
            caminho_destino = os.path.join(pasta_destino, arquivo)

            if os.path.isfile(caminho_origem):
                shutil.copy2(caminho_origem, caminho_destino)
                print(f"Arquivo {arquivo} copiado e substituído com sucesso!")
        
        print("Cópia concluída!")
    
    except Exception as e:
        print(f"Ocorreu um erro: {e}")


def atual_stockalerta():
    # Obter a data e hora atuais
    data_hora_atual = datetime.now()
    # Formatar a data e hora em uma string
    nomearquivo = data_hora_atual.strftime("%d-%m-%y %H%M%S")
    # Exibir a data e hora formatadas
    print("Data e hora atuais formatadas:", nomearquivo)
    fechar_outlook()
    abrirsap()
    sleep(15)
    verificar_msg()
    sleep(15)
    logarsap()
    sleep(15)
    verificar_msg()
    sleep(15)
    abrirtrans()
    sleep(10)
    abrirme2n()
    sleep(10)
    export_me2nSA()
    sleep(10)
    sairtransacao()
    sleep(10)
    sairtransacao()
    sleep(10)
    print('Primeiro arquivo gerado! Transação: ME2N')

    abrirtrans()
    sleep(10)
    abrirmb52()
    sleep(10)
    export_menu()
    sleep(10)
    export_mb52SA()
    sleep(10)
    sairtransacao()
    sleep(10)
    sairtransacao()
    print('Segundo arquivo gerado! Transação: MB52')

    abrirtrans()
    sleep(10)
    abrirzmb52()
    sleep(10)
    export_zmb52SA()
    sleep(10)
    sairtransacao()
    sleep(10)
    sairtransacao()
    sleep(10)
    print('Terceiro arquivo gerado! Transação: ZMB52')

    fecharsap()
    sleep(10)
    atualização_SA()
    tempo = datetime.now() - data_hora_atual
    print("Finalizando em:", tempo)
    print('**************BD Stock Alerta realizado com sucesso!*************')

def atualizacao():
    # Obter a data e hora atuais
    data_hora_atual = datetime.now()
    # Formatar a data e hora em uma string
    nomearquivo = data_hora_atual.strftime("%d-%m-%y %H%M%S")
    # Exibir a data e hora formatadas
    print("Data e hora atuais formatadas:", nomearquivo)
    fechar_outlook()
    abrirsap()
    sleep(2)
    verificar_msg()
    sleep(2)
    logarsap()
    sleep(2)
    verificar_msg()
    sleep(2)
    abrirjanelas()
    sleep(5)
    focus_window('ENP(1)/009 SAP Easy Access')
    abrircoois()
    variantepa()
    abrirrelat()
    clicarno()
    print('Iniciado COOIS PA')
    sleep(3)
    focus_window('ENP(2)/009 SAP Easy Access')
    sleep(2)
    trans_blank()
    print('Iniciado Blank')
    sleep(3)
    focus_window('ENP(3)/009 SAP Easy Access')
    sleep(2)
    abrircoois()
    variantepp()
    abrirrelat()
    exportexcel()
    salvarexcelpp()
    sairtransacao()
    sairtransacao()
    print('Primeiro arquivo gerado! Transação: COOIS')
    abririq09()
    exportexceliq09()
    sairtransacao()
    sairtransacao()
    print('Segundo arquivo gerado! Transação: IQ09')
    pyautogui.hotkey('alt','f4')
    sleep(15)
    export_menu()
    export_blank()
    print('Terceiro arquivo gerado! Transação: Blank')
    pyautogui.hotkey('alt','f4')
    sleep(30)
    controle_corte()
    sleep(20)
    exportexcel()
    salvarexcelpa()
    sairtransacao()
    sairtransacao()
    print('Quarto arquivo gerado! Transação: COOIS')
    sleep(2)
    fecharsap()
    data = pd.read_excel(r'O:\02-Blades\17-Warehouse-Operators\Levantamento de Material em Armazém\Material Adicional\EXCEL\Historico.xlsm', index_col=0)
    data.to_excel(r"O:\02-Blades\17-Warehouse-Operators\Levantamento de Material em Armazém\Material Adicional\EXCEL\historico_base.xlsx")
    atualização_BD()
    atualizarfarol_PP()
    tempo = datetime.now() - data_hora_atual
    print("Finalizando em:", tempo)
    print('**************BD Relatório Status de Produção atualizado com sucesso!*************')

def atualizar_backup():
    # Obter a data e hora atuais
    data_hora_atual = datetime.now()
    # Formatar a data e hora em uma string
    nomearquivo = data_hora_atual.strftime("%d-%m-%y %H%M%S")
    # Exibir a data e hora formatadas
    print("Data e hora atuais formatadas:", nomearquivo)
    fechar_outlook()

    # Obter a data de inicio e fim do periodo de pesquisa
    data_inicio = datetime.now() - timedelta(days=35)
    data_fim = datetime.now() + timedelta(days=35)
    # Formatar a data e hora em uma string
    datainicio = data_inicio.strftime("%d%m%Y")
    datafim = data_fim.strftime("%d%m%Y")

    abrirrescheduling()
    sleep(5)
    exerescheduling()
    sleep(5)
    export_backupr()
    sleep(5)
    print('Backup realizado com sucesso - Rescheduling')
    sleep(5)

    abrirbackupcoois()
    sleep(5)
    execoois()
    sleep(5)
    clicarno()
    sleep(30)
    export_backupcoois()
    sleep(5)
    print('Backup realizado com sucesso - Coois (ordens planeadas/ordens convertidas)')
    
    sleep(5)
    abrirbackupzmb52()
    sleep(5)
    exezmb52()
    sleep(5)
    export_backupzmb52()
    sleep(5)
    print('Backup realizado com sucesso - ZMB52')
    sleep(5)

    abrirbackupzbom()
    sleep(5)
    exezbom_e115()
    sleep(5)
    export_menu()
    sleep(5)
    export_backupzbom_e115()
    sleep(5)
    print('Backup realizado com sucesso - ZBOM E115')
    sleep(5)

    exezbom_e175()
    sleep(5)
    export_menu()
    sleep(5)
    export_backupzbom_e175()
    sleep(5)
    print('Backup realizado com sucesso - ZBOM E175')
    sleep(5)
    
    exezbom_e103()
    sleep(5)
    export_menu()
    sleep(5)
    export_backupzbom_e103()
    sleep(5)
    print('Backup realizado com sucesso - ZBOM E103')
    sleep(5)

    abrirbackupp2p()
    sleep(5)
    exep2p()
    sleep(5)
    export_backup2p()
    sleep(5)
    print('Backup realizado com sucesso - Mapa responsabilidade')
    sleep(5)
    pasta_origem = r'O:\02-Blades\02-Process Engineering\9. Projetos\22. Backup diario'
    pasta_destino = r'C:\Users\00082300\ENERCON\PT ROTO BCM Share - Priority\Planeamento da produção'
    copiar_arquivos(pasta_origem, pasta_destino)

    pasta_origem = r'O:\02-Blades\04-Production\04 - Production Planning\2025'
    pasta_destino = r'C:\Users\00082300\ENERCON\PT ROTO BCM Share - Priority\Planeamento da produção\2025'
    copiar_arquivos(pasta_origem, pasta_destino)
    tempo = datetime.now() - data_hora_atual
    print("Finalizando em:", tempo)



def backup_log():
    pasta_origem = r'\\srv-pt3\groups\02-Blades\05-Warehouse\ARMAZÉM\Fábrica Sul\3. Folhas para Carrinhos - Matéria Prima\4. Material por Peça E103'
    pasta_destino = r'C:\Users\00082300\ENERCON\PT ROTO BCM Share - Priority\Logistica - Armazem\IT-Breakdown MPP'
    copiar_arquivos(pasta_origem, pasta_destino)
    print("Backup realizado com sucesso - 4. Material por Peça E103 - Fábrica Sul")
    pasta_origem = r'\\srv-pt3\groups\02-Blades\05-Warehouse\ARMAZÉM\Fábrica Sul\3. Folhas para Carrinhos - Matéria Prima\5. Material por peça E175'
    pasta_destino = r'C:\Users\00082300\ENERCON\PT ROTO BCM Share - Priority\Logistica - Armazem\IT-Breakdown MPP'
    copiar_arquivos(pasta_origem, pasta_destino)
    print("Backup realizado com sucesso - 5. Material por peça E175 - Fábrica Sul")
    pasta_origem = r'\\srv-pt3\groups\02-Blades\05-Warehouse\ARMAZÉM\Fábrica Norte\2. MPP\MPP 175'
    pasta_destino = r'C:\Users\00082300\ENERCON\PT ROTO BCM Share - Priority\Logistica - Armazem\IT-Breakdown MPP'
    copiar_arquivos(pasta_origem, pasta_destino)
    print("Backup realizado com sucesso - MPP 175 - Fábrica Norte")
    pasta_origem = r'\\srv-pt3\groups\02-Blades\05-Warehouse\ARMAZÉM\Fábrica Norte\2. MPP\MPP103'
    pasta_destino = r'C:\Users\00082300\ENERCON\PT ROTO BCM Share - Priority\Logistica - Armazem\IT-Breakdown MPP'
    copiar_arquivos(pasta_origem, pasta_destino)
    print("Backup realizado com sucesso - MPP103 - Fábrica Norte")

schedule.every().monday.at("05:10").do(backup_log)
schedule.every().tuesday.at("05:10").do(backup_log)
schedule.every().wednesday.at("05:10").do(backup_log)
schedule.every().thursday.at("05:10").do(backup_log)
schedule.every().friday.at("05:10").do(backup_log)

schedule.every().monday.at("05:30").do(atualizar_backup)
schedule.every().tuesday.at("05:30").do(atualizar_backup)
schedule.every().wednesday.at("05:30").do(atualizar_backup)
schedule.every().thursday.at("05:30").do(atualizar_backup)
schedule.every().friday.at("05:30").do(atualizar_backup)
schedule.every().saturday.at("05:30").do(atualizar_backup)

schedule.every().monday.at("18:30").do(backup_log)
schedule.every().tuesday.at("18:30").do(backup_log)
schedule.every().wednesday.at("18:30").do(backup_log)
schedule.every().thursday.at("18:30").do(backup_log)
schedule.every().friday.at("18:30").do(backup_log)

schedule.every().monday.at("18:00").do(atualizar_backup)
schedule.every().tuesday.at("18:00").do(atualizar_backup)
schedule.every().wednesday.at("18:00").do(atualizar_backup)
schedule.every().thursday.at("18:00").do(atualizar_backup)
schedule.every().friday.at("18:00").do(atualizar_backup)
schedule.every().saturday.at("18:00").do(atualizar_backup)

schedule.every().monday.at("06:30").do(atual_stockalerta)
schedule.every().tuesday.at("06:30").do(atual_stockalerta)
schedule.every().wednesday.at("06:30").do(atual_stockalerta)
schedule.every().thursday.at("06:30").do(atual_stockalerta)
schedule.every().friday.at("06:30").do(atual_stockalerta)

schedule.every().monday.at("07:40").do(atualizacao)
schedule.every().tuesday.at("07:40").do(atualizacao)
schedule.every().wednesday.at("07:40").do(atualizacao)
schedule.every().thursday.at("07:40").do(atualizacao)
schedule.every().friday.at("07:40").do(atualizacao)

schedule.every().monday.at("15:40").do(atualizacao)
schedule.every().tuesday.at("15:40").do(atualizacao)
schedule.every().wednesday.at("15:40").do(atualizacao)
schedule.every().thursday.at("15:40").do(atualizacao)
schedule.every().friday.at("15:40").do(atualizacao)

while True:
    schedule.run_pending()
    sleep(30)