#-*- coding: UTF-8 -*-
import os
from time import sleep, time
import pyautogui
import pygetwindow as gw
import openpyxl
import schedule
import pyperclip
from datetime import datetime, timedelta
import shutil
import pandas as pd
import psutil
from openpyxl import load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from flask import Flask, jsonify, render_template
from threading import Timer

blank = 'blank'
pa = 'all_PA'
pp = 'all_PP'
sa1 = 'stock_alerta'
sa2 = 'stock_alerta2'
sa3 = 'stock_alerta3'
mb52 = 'mb52'
me2n = 'me2n'
zmb52 ='zmb52'
backupr ="backup_resched"
backupcoois = "backupcoois"
backupzmb52 = "backupzmb52"
e115_bom = "E115_BOM"
e103_bom = "E103_BOM"
e175_bom = "E175_BOM"
p2p = "scm_p2p"
scm = "scm_verificaca"

def fechar_outlook():
    # Verifica todos os processos ativos
    for proc in psutil.process_iter(['pid', 'name']):
        # Se o nome do processo for 'OUTLOOK.EXE', fecha-o
        if proc.info['name'].lower() == 'outlook.exe':
            print(f"Fechando Outlook (PID: {proc.info['pid']})...")
            proc.terminate()  # Tenta fechar o processo de forma educada
            try:
                proc.wait(timeout=3)  # Espera por até 3 segundos para o processo fechar
                print("Outlook fechado com sucesso.")
            except psutil.TimeoutExpired:
                print("O Outlook não fechou dentro do tempo esperado. Forçando o fechamento.")
                proc.kill()  # Força o fechamento se não fechar dentro do tempo
            return
    print("Outlook não está aberto.")
def check_message_on_screen(image_path, confidence=0.8):
    try:
        location = pyautogui.locateOnScreen(image_path, confidence=confidence)
        if location is not None:
            return True
        else:
            return False
    except pyautogui.ImageNotFoundException:
        return False
def verificar_msg():
    if __name__ == "__main__":
        image_path = r'C:\Users\00082300\Downloads\Projeto 2\export_conf1.png'  
        if check_message_on_screen(image_path):
            export_conf1 = pyautogui.locateCenterOnScreen('export_conf1.png', confidence=0.9)
            pyautogui.click(export_conf1[0],export_conf1[1], duration=1)
            print("A mensagem foi encontrada na tela.")
            sleep(10)
def enviar_erro(mensagem):
    # Função para enviar a mensagem de erro (implemente de acordo com sua necessidade)
    print(f"Erro: {mensagem}")

def esperar_e_clicar(imagem, descricao="Elemento", timeout=80, clicks=1, duracao=1, intervalo=1):
    """
    Aguarda até um tempo limite para a imagem aparecer na tela e clica nela.
    
    :param imagem: Caminho da imagem que deseja encontrar.
    :param descricao: Nome do elemento para exibição no console.
    :param timeout: Tempo máximo de espera antes de desistir.
    :param clicks: Número de cliques desejado.
    :param duracao: Tempo de duração do clique.
    :param intervalo: Tempo entre tentativas.
    :return: True se encontrou e clicou, False se o tempo acabar.
    """
    tempo_inicial = time()
    tentativas = 0
    
    while time() - tempo_inicial < timeout:
        try:
            localizacao = pyautogui.locateCenterOnScreen(imagem, confidence=0.9)
            
            if localizacao:
                print(f"{descricao} encontrado após {tentativas} tentativas! Clicando em {localizacao}")
                pyautogui.click(localizacao, duration=duracao, clicks=clicks)
                return True
        
        except pyautogui.ImageNotFoundException:
            print(f"Erro: '{descricao}' não encontrado na tentativa {tentativas}. Retentando...")

        tentativas += 1
        print(f"Tentativa {tentativas}/{timeout}: {descricao} não encontrado. Tentando novamente em {intervalo} segundos...")
        sleep(intervalo)

    print(f"Erro: Tempo esgotado! {descricao} não apareceu na tela após {timeout} segundos.")
    return False

def abrirsap():
    # ALL_PP
    #abrir o SAP
    os.startfile(r"C:\\Program Files (x86)\SAP\\FrontEnd\SAPgui\saplogon.exe")
    sleep(6)
def logarsap():
    esperar_e_clicar('botao_login_SAP.png', "Botão de Login SAP", timeout=80, clicks=2)

def abrircoois():
    #Abrir transação COOIS
    esperar_e_clicar('abrir_coois.png', "Abrir Coois", timeout=80, clicks=2)       
    

def variantepp():
    #abrir variante(pp)
    pyautogui.hotkey('shift', 'f5')
    sleep(4)
    esperar_e_clicar('dig_variavel.png', "Digitar Variante", timeout=80, clicks=2)    
    pyautogui.write(pp)
    esperar_e_clicar('conf_variavel.png', "Confirmar Variante", timeout=30, clicks=2)       

def abrirrelat():
    #Abrir relatório
    pyautogui.hotkey('f8')
    sleep(15)
def exportexcel():
    #abrir menu export
    esperar_e_clicar('abrir_menu_export.png', "Abrir Menu Exportar", timeout=80, clicks=2)       
    
    #Exportar para excel
    imagem = 'export_excel.png'
    esperar_e_clicar('export_excel.png', "Confirmar Botão Export", timeout=30, clicks=2)       

    imagem = 'export_excel_menu.png'
    esperar_e_clicar('export_excel_menu.png', "Export Menu Excel", timeout=30, clicks=2)       

    imagem = 'conf_export.png'
    esperar_e_clicar('conf_export.png', "Confirmar Export", timeout=30, clicks=2)       

def salvarexcelpp():
    #Salvar a planilha
    all_pp_salve = r'\\srv-pt3\groups\02-Blades\02-Process Engineering\9. Projetos\4. Dashboard\Base de Dados\all_pp.XLSX'
    pyautogui.write(all_pp_salve)
    pyautogui.hotkey('enter')
    sleep(4)
    pyautogui.hotkey('y')
    sleep(4)
    pyautogui.hotkey('alt','f4')
    sleep(2)
def sairtransacao():
    #sair transação
    pyautogui.hotkey('esc')
    sleep(3)
def variantepa():
    #abrir variante (PA)
    pyautogui.hotkey('shift', 'f5')
    esperar_e_clicar('dig_variavel.png', "Digitar Variante", timeout=30, clicks=2)       
    pyautogui.write(pa)
    esperar_e_clicar('conf_variavel.png', "Confirmar Variante", timeout=30, clicks=2)       

def clicarno():
    #clicar no botão no para muitas linhas(superior a 5mil linhas)
    esperar_e_clicar('botao_no.png', "Clicar no Botão Não", timeout=30, clicks=2)       
    
def salvarexcelpa():
    #Salvar a planilha
    sleep(90)
    all_pa_salve = r'\\srv-pt3\groups\02-Blades\02-Process Engineering\9. Projetos\4. Dashboard\Base de Dados\All - Por Ano\All_2025.XLSX'
    pyautogui.write(all_pa_salve)
    sleep(10)
    pyautogui.hotkey('enter')
    sleep(10)
    pyautogui.hotkey('y')
    sleep(120)
    pyautogui.hotkey('alt','f4')
    sleep(10)
def abririq09():
    #Abrir transação IQ09 Stock
    esperar_e_clicar('abrir_iq09.png', "Abrir IQ09", timeout=30, clicks=2)       
    
    #abrir variante(pp)
    pyautogui.hotkey('shift', 'f5')
    esperar_e_clicar('dig_variavel.png', "Digitar Variante", timeout=30, clicks=2)       
    pyautogui.write(pp)
    esperar_e_clicar('conf_variavel.png', "Confirmar Variante", timeout=30, clicks=2)       
    pyautogui.hotkey('f8')
    sleep(4)
def exportexceliq09():
    #Exportar excel
    esperar_e_clicar('export_iq09.png', "Export IQ 09", timeout=30, clicks=2)       
        
    esperar_e_clicar('export_conf1.png', "Confirmar Exporte 09", timeout=30, clicks=2)       
    pyautogui.hotkey('up')
    sleep(2)
    esperar_e_clicar('export_conf1.png', "Confirmar Exporte 09", timeout=30, clicks=2)       
    
    esperar_e_clicar('export_conf1.png', "Confirmar Exporte 09", timeout=30, clicks=2)       
    
    esperar_e_clicar('salve_table.png', "Salvar Tabela", timeout=30, clicks=2)       
    pyautogui.hotkey('f12')
    sleep(3)
    all_pa_salve = r'\\srv-pt3\groups\02-Blades\02-Process Engineering\9. Projetos\4. Dashboard\Base de Dados\Stock_all_pp.xlsx'
    pyautogui.hotkey('backspace')
    sleep(3)
    pyautogui.hotkey('alt','n')
    sleep(3)
    pyautogui.write(all_pa_salve)
    pyautogui.hotkey('enter')
    sleep(1)
    pyautogui.hotkey('left')
    pyautogui.hotkey('enter')
    sleep(5)
    pyautogui.hotkey('alt','f4')
    sleep(1)
def fecharsap():
    #sair transação
    pyautogui.hotkey('alt','f4')
    sleep(1)

    esperar_e_clicar('yes_sair.png', "clicar em sim", timeout=30, clicks=2)    
    #fechar o SAP
    esperar_e_clicar('sair_x.png', "Fechar SAP", timeout=30, clicks=2)    
    
def abrirtrans():
    #Abrir transação COOIS
    esperar_e_clicar('digitar_trans.png', "Fechar SAP", timeout=30, clicks=2)    

def abrirmb52():
    pyautogui.write(mb52)
    pyautogui.hotkey('enter')
    sleep(2)
    #abrir variante(sa1)
    pyautogui.hotkey('shift', 'f5')
    sleep(4)
    esperar_e_clicar('dig_variavel.png', "Digitar Variavel", timeout=30, clicks=1)  

    pyautogui.write(sa1)
    esperar_e_clicar('conf_variavel.png', "Confirmar Variavel", timeout=30, clicks=2)  
    pyautogui.hotkey('f8')
    sleep(4)
def abrirme2n():
    pyautogui.write(me2n)
    pyautogui.hotkey('enter')
    sleep(2)
    #abrir variante(sa1)
    pyautogui.hotkey('shift', 'f5')
    sleep(4)
    esperar_e_clicar('dig_variavel.png', "Confirmar Variavel", timeout=30, clicks=1)  
    
    pyautogui.write(sa2)
    
    esperar_e_clicar('conf_variavel.png', "Confirmar Variavel", timeout=30, clicks=1)  

    esperar_e_clicar('acesso_variante.png', "Confirmar Variavel", timeout=30, clicks=1)  

    esperar_e_clicar('Conf_alteracao.png', "Confirmar Variavel", timeout=30, clicks=1) 
    pyautogui.hotkey('f8')
    sleep(30)
def abrirzmb52():
    pyautogui.write(zmb52)
    pyautogui.hotkey('enter')
    sleep(2)
    #abrir variante(sa1)
    pyautogui.hotkey('shift', 'f5')
    sleep(4)

    esperar_e_clicar('dig_variavel.png', "Confirmar Variavel", timeout=30, clicks=1)

    pyautogui.write(sa3)

    esperar_e_clicar('conf_variavel.png', "Confirmar Variavel", timeout=30, clicks=1)

def trans_blank():
    #abrir transação
    esperar_e_clicar('mb51_serial.png', "Confirmar Variavel", timeout=30, clicks=2)
    #abrir variante(blank)
    pyautogui.hotkey('shift', 'f5')
    sleep(4)

    esperar_e_clicar('janelaexportar.png', "Confirmar Variavel", timeout=30, clicks=2)

    pyautogui.write(blank)
    esperar_e_clicar('conf_variavel.png', "Confirmar Variavel", timeout=30, clicks=2)
    pyautogui.hotkey('f8')
    sleep(10)
def export_menu():
    esperar_e_clicar('menu-list.png', "Confirmar Variavel", timeout=30, clicks=1)

    esperar_e_clicar('menu-export.png', "Confirmar Variavel", timeout=30, clicks=1)

    esperar_e_clicar('menu_sheet.png', "Confirmar Variavel", timeout=30, clicks=1)

    esperar_e_clicar('conf_export.png', "Confirmar Variavel", timeout=30, clicks=1)
    sleep(30)
def export_blank():
    #Salvar a planilha
    all_blank = r'\\srv-pt3\groups\02-Blades\02-Process Engineering\9. Projetos\4. Dashboard\Base de Dados\Serialnumber_blank.xlsx'
    pyautogui.write(all_blank)
    pyautogui.hotkey('enter')
    sleep(10)
    pyautogui.hotkey('y')
    sleep(20)
    pyautogui.hotkey('alt','f4')
    sleep(2)
def export_mb52SA():
    #Salvar a planilha
    mb52SA = r'\\srv-pt3\groups\02-Blades\05-Warehouse\08. Análise semanal de níveis de stock\Base de Dados\SAP\MB52_All.XLSX'
    pyperclip.copy(mb52SA)
    pyautogui.hotkey('ctrl', 'v')
    pyautogui.hotkey('enter')
    sleep(4)
    pyautogui.hotkey('y')
    sleep(60)
    pyautogui.hotkey('alt','f4')
    sleep(2)
def export_me2nSA():
    esperar_e_clicar('abrirdatas.png', "Confirmar Variavel", timeout=30, clicks=2)
    esperar_e_clicar('devildate.png', "Confirmar Variavel", timeout=30, clicks=1)

    esperar_e_clicar('acendente.png', "Confirmar Variavel", timeout=30, clicks=1)

    esperar_e_clicar('export_iq09.png', "Confirmar Variavel", timeout=30, clicks=1)
    esperar_e_clicar('conf_export.png', "Confirmar Variavel", timeout=30, clicks=1)

    #Salvar a planilha
    me2nSA = r'\\srv-pt3\groups\02-Blades\05-Warehouse\08. Análise semanal de níveis de stock\Base de Dados\SAP\ME2N_All.XLSX'
    pyperclip.copy(me2nSA)
    pyautogui.hotkey('ctrl', 'v')
    pyautogui.hotkey('enter')
    sleep(4)
    pyautogui.hotkey('y')
    sleep(60)
    pyautogui.hotkey('alt','f4')
    sleep(2)
def export_zmb52SA():
    pyautogui.hotkey('ctrl','shift','f7')
    sleep(2)

    esperar_e_clicar('conf_export.png', "Confirmar Variavel", timeout=30, clicks=1)
    #Salvar a planilha
    zmb52SA = r'\\srv-pt3\groups\02-Blades\05-Warehouse\08. Análise semanal de níveis de stock\Base de Dados\SAP\ZMB52_All.XLSX'
    pyperclip.copy(zmb52SA)
    pyautogui.hotkey('ctrl', 'v')
    pyautogui.hotkey('enter')
    sleep(4)
    pyautogui.hotkey('y')
    sleep(60)
    pyautogui.hotkey('alt','f4')
    sleep(2)
def controle_corte():
    os.startfile(r"\\srv-pt3\groups\02-Blades\02-Process Engineering\9. Projetos\4. Dashboard\Base de Dados\CONTROLE CORTE.xlsm")
    sleep(45)
    pyautogui.hotkey('alt', 'f4')
    sleep(4)
    pyautogui.hotkey('g')
    print('Controle de Corte atualizado!')
    sleep(10)
def atualização_BD():
    
    os.startfile(r'\\srv-pt3\groups\02-Blades\02-Process Engineering\9. Projetos\4. Dashboard\Base de Dados\BD_2024_dashboard.xlsm')
    sleep(30)
    pyautogui.hotkey('ctrl', 'alt','f5')
    pyautogui.hotkey('left')
    pyautogui.hotkey('enter')
    print('Atualizando powerquery')
    sleep(120)
    print('Finalizado o prazo de atualização PowerQuery')
    pyautogui.hotkey('ctrl', 'q')
    print('Atualizando Macro')
    sleep(250)
    print('Finalziado a atualização da Macro')
def abrirjanelas():
    #abrir 3 janelas
    sleep(5)
    pyautogui.hotkey('ctrl', 'n')
    sleep(5)
    pyautogui.hotkey('ctrl', 'n')
    sleep(5)
def focus_window(title):
    windows = gw.getWindowsWithTitle(title)
    if windows:
        window = windows[0]
        window.activate()
    else:
        print(f'Janela com título "{title}" não encontrada.')
def atualização_SA():
    os.startfile(r'\\srv-pt3\groups\02-Blades\05-Warehouse\08. Análise semanal de níveis de stock\Base de Dados\Teste\2. PTA3_stock alerta(MACRO).xlsm')
    sleep(15)
    pyautogui.hotkey('ctrl', 'w')
    print('Atualizando Macro')
    sleep(120)
    pyautogui.hotkey('enter')
    print('Finalziado a atualização da Macro PTA3')
    pyautogui.hotkey('alt', 'f4')
    sleep(4)
    pyautogui.hotkey('g')
    print('Finalziado arquivo PTA3')
    sleep(4)
    os.startfile(r'\\srv-pt3\groups\02-Blades\05-Warehouse\08. Análise semanal de níveis de stock\Base de Dados\Teste\4. STOCK ALERTA PTA0 _PTA3.xlsm')
    sleep(15)
    pyautogui.hotkey('ctrl', 'w')
    print('Atualizando Macro')
    sleep(120)
    pyautogui.hotkey('enter')
    print('Finalziado a atualização da Macro PTA0')
    pyautogui.hotkey('alt', 'f4')
    sleep(4)
    pyautogui.hotkey('g')
    print('Finalziado arquivo PTA0')

def atualizarfarol_PP():
    # Obter a data e hora atuais
    data_hora_atual = datetime.now()
    # Formatar a data e hora em uma string
    nomearquivo = data_hora_atual.strftime("%d-%m-%y %H%M%S")
    # Exibir a data e hora formatadas
    print("Data e hora atuais formatadas:", nomearquivo)
    # Caminho da pasta com os arquivos Excel
    pasta = r"\\srv-pt3\groups\02-Blades\02-Process Engineering\9. Projetos\4. Dashboard\Base de Dados\All - Por Ano"
    # Lista todos os arquivos dentro da pasta
    arquivos = [os.path.join(pasta, arquivo) for arquivo in os.listdir(pasta)]

    # Inicializa um DataFrame vazio para concatenar os dados
    tabela_final = pd.DataFrame()

    # Nome da coluna que será filtrada e a lista de valores desejados para o filtro
    coluna_filtro = "Work Center"
    valores_desejados = ["PP_M005","QA0210","QA0232","PP_M020","PP_M018","QA0230","PP_M019","QA0231","PP_M017","QA0229","PP_M006","QA0211","PP_M013",
    "QA0226","PP_M014","QA0227","PP_M015","QA0228","PP_M011","QA0221","PP_PA_N","QA0222","QA0223","PP_M004","QA0208","MM_PA_S","QA0209","PP_M008","QA0214","QA0215","PP_M012","QA0224","QA0225","PP_M007","QA0212","QA0213","PP_M010","QA0218","QA0219","QA0220","PP_M003","QA0206","QA0207","PP_M002","QA0203","QA0204","QA0205","PP_M009","QA0216","QA0217","PP_M001","QA0200","QA0201","QA0202"]  # Lista com os valores desejados

    for arquivo in arquivos:
        df = pd.read_excel(arquivo, index_col=0)  # Lê o arquivo Excel
        df_filtrado = df[df[coluna_filtro].isin(valores_desejados)]  # Filtra os dados que possuem valores da lista
        tabela_final = pd.concat([tabela_final, df_filtrado])  # Concatena os dados filtrados na tabela final

    # Exporta a tabela final concatenada e filtrada para um arquivo Excel

    tabela_final.to_excel(r"\\srv-pt3\groups\02-Blades\04-Production\01 - Raw Parts\29 - Primary Parts\01 - Produção\09 - Farol MES\BD\BD_PP.xlsx")
    print("Importação, filtro e concatenização concluídos com sucesso!")

    df = pd.DataFrame(tabela_final)

    # Exportar para Excel (criando o arquivo inicial)
    arquivo_excel = r'\\srv-pt3\groups\02-Blades\04-Production\01 - Raw Parts\29 - Primary Parts\01 - Produção\09 - Farol MES\BD\BD_PP.xlsx'
    df.to_excel(arquivo_excel, index=False, sheet_name='Planilha1')

    # Abrir o arquivo com openpyxl para formatar
    workbook = load_workbook(arquivo_excel)
    planilha = workbook['Planilha1']

    # Definir o intervalo dos dados como uma tabela
    # Intervalo automático com base nos dados exportados
    inicio_celula = planilha.cell(row=1, column=1).coordinate  # A1
    fim_celula = planilha.cell(row=1 + len(df), column=len(df.columns)).coordinate  # Última célula

    # Criar a tabela
    tabela = Table(displayName="Tabela1", ref=f"{inicio_celula}:{fim_celula}")

    # Definir o estilo da tabela
    estilo = TableStyleInfo(
        name="TableStyleMedium9",  # Escolha o estilo desejado
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=True
    )
    tabela.tableStyleInfo = estilo

    # Adicionar a tabela à planilha
    planilha.add_table(tabela)

    # Salvar o arquivo Excel com a tabela formatada
    workbook.save(arquivo_excel)
    workbook.close()

    print(f"Arquivo Excel '{arquivo_excel}' criado e formatado com sucesso!")
    tempo = datetime.now() - data_hora_atual
    print("Finalizando em:", tempo)

def atualizar_bd_PP():
        # Obter a data e hora atuais
    data_hora_atual = datetime.now()
    # Formatar a data e hora em uma string
    nomearquivo = data_hora_atual.strftime("%d-%m-%y %H%M%S")
    # Exibir a data e hora formatadas
    print("Data e hora atuais formatadas:", nomearquivo)
    fechar_outlook()
    abrirsap()
    sleep(2)
    verificar_msg()
    sleep(2)
    logarsap()
    sleep(2)
    verificar_msg()
    sleep(2)
    abrircoois()
    variantepa()
    abrirrelat()
    clicarno()
    print('Iniciado COOIS PA')
    sleep(350)
    controle_corte()
    sleep(20)
    exportexcel()
    salvarexcelpa()
    sairtransacao()
    sairtransacao()
    print('Arquivo gerado! Transação: COOIS')
    sleep(2)
    fecharsap()
    atualizarfarol_PP()
    tempo = datetime.now() - data_hora_atual
    print("Finalizando em:", tempo)
    print('**************BD Relatório Status de Produção atualizado com sucesso!*************')

def abrirrescheduling():
    # Abrir o aplicativo
    os.startfile(r"C:\Users\00082300\Downloads\Projeto 2\atalhos\mm_utr.SAP")
    sleep(10)

    # Tentar encontrar a janela aberta
    windows = gw.getWindowsWithTitle('ENP')

    if windows:
        app_window = windows[0]
        app_window.maximize()
    else:
        print("Janela do aplicativo não encontrada")
def abrirbackupcoois():
    # Abrir o aplicativo
    os.startfile(r"C:\Users\00082300\Downloads\Projeto 2\atalhos\coois.SAP")
    sleep(10)

    # Tentar encontrar a janela aberta
    windows = gw.getWindowsWithTitle('ENP')

    if windows:
        app_window = windows[0]
        app_window.maximize()
    else:
        print("Janela do aplicativo não encontrada")   
def abrirbackupzmb52():
    # Abrir o aplicativo
    os.startfile(r"C:\Users\00082300\Downloads\Projeto 2\atalhos\zmb52.SAP")
    sleep(10)

    # Tentar encontrar a janela aberta
    windows = gw.getWindowsWithTitle('ENP')

    if windows:
        app_window = windows[0]
        app_window.maximize()
    else:
        print("Janela do aplicativo não encontrada")  
def abrirbackupzbom():
    # Abrir o aplicativo
    os.startfile(r"C:\Users\00082300\Downloads\Projeto 2\atalhos\zbom.SAP")
    sleep(10)

    # Tentar encontrar a janela aberta
    windows = gw.getWindowsWithTitle('ENP')

    if windows:
        app_window = windows[0]
        app_window.maximize()
    else:
        print("Janela do aplicativo não encontrada")  
def abrirbackupp2p():
    # Abrir o aplicativo
    os.startfile(r"C:\Users\00082300\Downloads\Projeto 2\atalhos\p2p.SAP")
    sleep(10)

    # Tentar encontrar a janela aberta
    windows = gw.getWindowsWithTitle('ENP')

    if windows:
        app_window = windows[0]
        app_window.maximize()
    else:
        print("Janela do aplicativo não encontrada")  

def exerescheduling():
    pyautogui.hotkey('shift', 'f5')
    sleep(4)

    esperar_e_clicar('janelaexportar.png', "Confirmar Variavel", timeout=30, clicks=2)
    pyautogui.write(backupr)

    esperar_e_clicar('conf_variavel.png', "Confirmar Variavel", timeout=30, clicks=2)

    pyautogui.hotkey('f8')
    sleep(80)
def execoois():
    # Obter a data e hora atuais
    data_hora_atual = datetime.now()
    # Formatar a data e hora em uma string
    nomearquivo = data_hora_atual.strftime("%d-%m-%y %H%M%S")

    # Obter a data de inicio e fim do periodo de pesquisa
    data_inicio = datetime.now() - timedelta(days=35)
    data_fim = datetime.now() + timedelta(days=35)
    # Formatar a data e hora em uma string
    datainicio = data_inicio.strftime("%d%m%Y")
    datafim = data_fim.strftime("%d%m%Y")
    pyautogui.hotkey('shift', 'f5')
    sleep(4)

    esperar_e_clicar('janelaexportar.png', "Confirmar Variavel", timeout=30, clicks=2)
    sleep(4)
    pyautogui.write(backupcoois)

    esperar_e_clicar('conf_variavel.png', "Confirmar Variavel", timeout=30, clicks=2)
    esperar_e_clicar('datestart.png', "Confirmar Variavel", timeout=30, clicks=2)
    esperar_e_clicar('X_vemerlho.png', "Confirmar Variavel", timeout=30, clicks=1)
    sleep(6)
    pyautogui.write(datainicio)

    esperar_e_clicar('datefinish.png', "Confirmar Variavel", timeout=30, clicks=2)
    esperar_e_clicar('X_vemerlho.png', "Confirmar Variavel", timeout=30, clicks=1)
    pyautogui.write(datafim)

    pyautogui.hotkey('f8')
    sleep(80)
def exezmb52():
    pyautogui.hotkey('shift', 'f5')
    sleep(4)

    esperar_e_clicar('janelaexportar.png', "Confirmar Variavel", timeout=30, clicks=2)
    sleep(4)
    pyautogui.write(backupzmb52)

    esperar_e_clicar('conf_variavel.png', "Confirmar Variavel", timeout=30, clicks=2)
    sleep(4)
    pyautogui.hotkey('f8')
    sleep(80)
def exep2p():
    pyautogui.hotkey('shift', 'f5')
    sleep(4)

    esperar_e_clicar('janelaexportar.png', "Confirmar Variavel", timeout=30, clicks=2)
    sleep(4)
    pyautogui.write(p2p)

    esperar_e_clicar('conf_variavel.png', "Confirmar Variavel", timeout=30, clicks=2)
    pyautogui.hotkey('f8')
    sleep(80)

def exezbom_e115():
    pyautogui.hotkey('shift', 'f5')
    sleep(4)

    esperar_e_clicar('janelaexportar.png', "Confirmar Variavel", timeout=30, clicks=2)
    sleep(4)
    pyautogui.write(e115_bom)

    esperar_e_clicar('conf_variavel.png', "Confirmar Variavel", timeout=30, clicks=2)
    sleep(6)
    pyautogui.hotkey('f8')
    sleep(80)
def exezbom_e175():
    pyautogui.hotkey('shift', 'f5')
    sleep(4)
    esperar_e_clicar('janelaexportar.png', "Confirmar Variavel", timeout=30, clicks=2)
    sleep(4)
    pyautogui.write(e175_bom)

    esperar_e_clicar('conf_variavel.png', "Confirmar Variavel", timeout=30, clicks=2)
    sleep(6)
    pyautogui.hotkey('f8')
    sleep(80)
def exezbom_e103():
    pyautogui.hotkey('shift', 'f5')
    sleep(4)

    esperar_e_clicar('janelaexportar.png', "Confirmar Variavel", timeout=30, clicks=2)
    sleep(4)
    pyautogui.write(e103_bom)

    esperar_e_clicar('conf_variavel.png', "Confirmar Variavel", timeout=30, clicks=2)
    sleep(4)
    pyautogui.hotkey('f8')
    sleep(80)

def export_backupr():
    pyautogui.hotkey('shift','f9')
    sleep(4)
    
    esperar_e_clicar('conf_export.png', "Confirmar Variavel", timeout=30, clicks=1)
    sleep(60)
    arquivo = r'O:\02-Blades\02-Process Engineering\9. Projetos\22. Backup diario\IT-Breakdown Open Purchase Orders Actual Status'
    pyautogui.write(arquivo)
    pyautogui.hotkey('enter')
    sleep(6)
    pyautogui.hotkey('y')
    sleep(60)
    pyautogui.hotkey('f12')
    arquivo = r'C:\Users\00082300\ENERCON\PT ROTO BCM Share - Priority\Planeamento da produção\IT-Breakdown Open Purchase Orders Actual Status'
    pyperclip.copy(arquivo)
    sleep(4)
    pyautogui.hotkey('ctrl', 'v')
    sleep(4)
    pyautogui.hotkey('enter')
    sleep(6)
    pyautogui.hotkey('s')
    sleep(60)
    pyautogui.hotkey('alt','f4')
    sleep(10)

    esperar_e_clicar('sair_x.png', "Confirmar Variavel", timeout=30, clicks=1)
    esperar_e_clicar('yes_sair.png', "Confirmar Variavel", timeout=30, clicks=1)

def export_backupcoois():
    #abrir menu export
    esperar_e_clicar('abrir_menu_export.png', "Confirmar Variavel", timeout=30, clicks=1)
    #Exportar para excel
    esperar_e_clicar('export_excel.png', "Confirmar Variavel", timeout=30, clicks=1)
    esperar_e_clicar('export_excel_menu.png', "Confirmar Variavel", timeout=30, clicks=1)
    esperar_e_clicar('conf_export.png', "Confirmar Variavel", timeout=30, clicks=1)
    sleep(60)
    arquivo = r'O:\02-Blades\02-Process Engineering\9. Projetos\22. Backup diario\IT-Breakdown Production Orders_Converted and Planned'
    pyautogui.write(arquivo)
    pyautogui.hotkey('enter')
    sleep(6)
    pyautogui.hotkey('y')
    sleep(60)
    pyautogui.hotkey('f12')
    arquivo = r'C:\Users\00082300\ENERCON\PT ROTO BCM Share - Priority\Planeamento da produção\IT-Breakdown Production Orders_Converted and Planned'
    pyperclip.copy(arquivo)
    sleep(4)
    pyautogui.hotkey('ctrl', 'v')
    sleep(4)
    pyautogui.hotkey('enter')
    sleep(6)
    pyautogui.hotkey('s')
    sleep(60)
    pyautogui.hotkey('alt','f4')
    sleep(10)
    
    esperar_e_clicar('sair_x.png', "Confirmar Variavel", timeout=30, clicks=2)
    sleep(4)
    esperar_e_clicar('yes_sair.png', "Confirmar Variavel", timeout=30, clicks=2)
    
def export_backupzmb52():
    pyautogui.hotkey('ctrl','shift','f7')
    sleep(4)
    esperar_e_clicar('conf_export.png', "Confirmar Variavel", timeout=30, clicks=1)
    sleep(60)
    arquivo = r'O:\02-Blades\02-Process Engineering\9. Projetos\22. Backup diario\IT-Breakdown Current Stocks PTA0_PTA3 '
    pyautogui.write(arquivo)
    pyautogui.hotkey('enter')
    sleep(6)
    pyautogui.hotkey('y')
    sleep(60)
    pyautogui.hotkey('f12')
    arquivo = r'C:\Users\00082300\ENERCON\PT ROTO BCM Share - Priority\Planeamento da produção\IT-Breakdown Current Stocks PTA0_PTA3 '
    pyperclip.copy(arquivo)
    sleep(4)
    pyautogui.hotkey('ctrl', 'v')
    sleep(4)
    pyautogui.hotkey('enter')
    sleep(6)
    pyautogui.hotkey('s')
    sleep(60)
    pyautogui.hotkey('alt','f4')
    sleep(10)

    esperar_e_clicar('sair_x.png', "Confirmar Variavel", timeout=30, clicks=1)
    esperar_e_clicar('yes_sair.png', "Confirmar Variavel", timeout=30, clicks=1)

def export_backup2p():
    #Exportar para excel
    esperar_e_clicar('export_excel.png', "Confirmar Variavel", timeout=30, clicks=1)
    esperar_e_clicar('export_excel_menu.png', "Confirmar Variavel", timeout=30, clicks=1)
    esperar_e_clicar('conf_export.png', "Confirmar Variavel", timeout=30, clicks=1)
    sleep(60)
    arquivo = r'O:\02-Blades\02-Process Engineering\9. Projetos\22. Backup diario\IT-Breakdown Global Material Responsibles'
    pyautogui.write(arquivo)
    pyautogui.hotkey('enter')
    sleep(6)
    pyautogui.hotkey('y')
    sleep(60)
    pyautogui.hotkey('f12')
    arquivo = r'C:\Users\00082300\ENERCON\PT ROTO BCM Share - Priority\Planeamento da produção\IT-Breakdown Global Material Responsibles'
    pyperclip.copy(arquivo)
    sleep(6)
    pyautogui.hotkey('ctrl', 'v')
    sleep(6)
    pyautogui.hotkey('enter')
    sleep(6)
    pyautogui.hotkey('s')
    sleep(60)
    pyautogui.hotkey('alt','f4')
    sleep(10)
    esperar_e_clicar('sair_x.png', "Confirmar Variavel", timeout=30, clicks=1)
    esperar_e_clicar('yes_sair.png', "Confirmar Variavel", timeout=30, clicks=1)

def export_backupzbom_e115():
    arquivo = r'O:\02-Blades\02-Process Engineering\9. Projetos\22. Backup diario\IT-Breakdown BOM_Standard Grey Blade E115'
    pyautogui.write(arquivo)
    pyautogui.hotkey('enter')
    sleep(4)
    pyautogui.hotkey('y')
    sleep(60)
    pyautogui.hotkey('f12')
    arquivo = r'C:\Users\00082300\ENERCON\PT ROTO BCM Share - Priority\Planeamento da produção\IT-Breakdown BOM_Standard Grey Blade E115'
    pyperclip.copy(arquivo)
    sleep(2)
    pyautogui.hotkey('ctrl', 'v')
    sleep(2)
    pyautogui.hotkey('enter')
    sleep(4)
    pyautogui.hotkey('s')
    sleep(60)
    pyautogui.hotkey('alt','f4')
    sleep(10)
    pyautogui.hotkey('esc')
    sleep(3)
def export_backupzbom_e175():
    arquivo = r'O:\02-Blades\02-Process Engineering\9. Projetos\22. Backup diario\IT-Breakdown BOM_Standard Grey Blade E175'
    pyautogui.write(arquivo)
    pyautogui.hotkey('enter')
    sleep(4)
    pyautogui.hotkey('y')
    sleep(60)
    pyautogui.hotkey('f12')
    arquivo = r'C:\Users\00082300\ENERCON\PT ROTO BCM Share - Priority\Planeamento da produção\IT-Breakdown BOM_Standard Grey Blade E175'
    pyperclip.copy(arquivo)
    sleep(4)
    pyautogui.hotkey('ctrl', 'v')
    sleep(4)
    pyautogui.hotkey('enter')
    sleep(6)
    pyautogui.hotkey('s')
    sleep(60)
    pyautogui.hotkey('alt','f4')
    sleep(10)
    pyautogui.hotkey('esc')
    sleep(4)
def export_backupzbom_e103():
    arquivo = r'O:\02-Blades\02-Process Engineering\9. Projetos\22. Backup diario\IT-Breakdown BOM_Standard Grey Blade E103'
    pyautogui.write(arquivo)
    pyautogui.hotkey('enter')
    sleep(6)
    pyautogui.hotkey('y')
    sleep(60)
    pyautogui.hotkey('f12')
    arquivo = r'C:\Users\00082300\ENERCON\PT ROTO BCM Share - Priority\Planeamento da produção\IT-Breakdown BOM_Standard Grey Blade E103'
    pyperclip.copy(arquivo)
    sleep(4)
    pyautogui.hotkey('ctrl', 'v')
    sleep(4)
    pyautogui.hotkey('enter')
    sleep(6)
    pyautogui.hotkey('s')
    sleep(60)
    pyautogui.hotkey('alt','f4')
    sleep(10)
    esperar_e_clicar('sair_x.png', "Confirmar Variavel", timeout=30, clicks=1)
    esperar_e_clicar('yes_sair.png', "Confirmar Variavel", timeout=30, clicks=1)

def copiar_arquivos(pasta_origem, pasta_destino):
    try:
        if not os.path.exists(pasta_destino):
            os.makedirs(pasta_destino)

        arquivos = os.listdir(pasta_origem)
        for arquivo in arquivos:
            caminho_origem = os.path.join(pasta_origem, arquivo)
            caminho_destino = os.path.join(pasta_destino, arquivo)

            if os.path.isfile(caminho_origem):
                shutil.copy2(caminho_origem, caminho_destino)
                print(f"Arquivo {arquivo} copiado e substituído com sucesso!")
        
        print("Cópia concluída!")
    
    except Exception as e:
        print(f"Ocorreu um erro: {e}")


#def atual_stockalerta():
# Obter a data e hora atuais
data_hora_atual = datetime.now()
# Formatar a data e hora em uma string
nomearquivo = data_hora_atual.strftime("%d-%m-%y %H%M%S")
# Exibir a data e hora formatadas
print("Data e hora atuais formatadas:", nomearquivo)
fechar_outlook()
abrirsap()
verificar_msg()
logarsap()
verificar_msg()
abrirtrans()
abrirme2n()
export_me2nSA()
sairtransacao()
sairtransacao()
print('Primeiro arquivo gerado! Transação: ME2N')

abrirtrans()
abrirmb52()
export_menu()
export_mb52SA()
sairtransacao()
sairtransacao()
print('Segundo arquivo gerado! Transação: MB52')

abrirtrans()
abrirzmb52()
export_zmb52SA()
sairtransacao()
sairtransacao()
print('Terceiro arquivo gerado! Transação: ZMB52')

fecharsap()
tempo = datetime.now() - data_hora_atual
print("Finalizando em:", tempo)
print('**************BD Stock Alerta realizado com sucesso!*************')

def atualizacao():
    # Obter a data e hora atuais
    data_hora_atual = datetime.now()
    # Formatar a data e hora em uma string
    nomearquivo = data_hora_atual.strftime("%d-%m-%y %H%M%S")
    # Exibir a data e hora formatadas
    print("Data e hora atuais formatadas:", nomearquivo)
    fechar_outlook()
    abrirsap()
    sleep(2)
    verificar_msg()
    sleep(2)
    logarsap()
    sleep(2)
    verificar_msg()
    sleep(2)
    abrirjanelas()
    sleep(5)
    focus_window('ENP(1)/009 SAP Easy Access')
    abrircoois()
    variantepa()
    abrirrelat()
    clicarno()
    print('Iniciado COOIS PA')
    sleep(3)
    focus_window('ENP(2)/009 SAP Easy Access')
    sleep(2)
    trans_blank()
    print('Iniciado Blank')
    sleep(3)
    focus_window('ENP(3)/009 SAP Easy Access')
    sleep(2)
    abrircoois()
    variantepp()
    abrirrelat()
    exportexcel()
    salvarexcelpp()
    sairtransacao()
    sairtransacao()
    print('Primeiro arquivo gerado! Transação: COOIS')
    abririq09()
    exportexceliq09()
    sairtransacao()
    sairtransacao()
    print('Segundo arquivo gerado! Transação: IQ09')
    pyautogui.hotkey('alt','f4')
    sleep(15)
    export_menu()
    export_blank()
    print('Terceiro arquivo gerado! Transação: Blank')
    pyautogui.hotkey('alt','f4')
    sleep(30)
    controle_corte()
    sleep(20)
    exportexcel()
    salvarexcelpa()
    sairtransacao()
    sairtransacao()
    print('Quarto arquivo gerado! Transação: COOIS')
    sleep(2)
    fecharsap()
    data = pd.read_excel(r'O:\02-Blades\17-Warehouse-Operators\Levantamento de Material em Armazém\Material Adicional\EXCEL\Historico.xlsm', index_col=0)
    data.to_excel(r"O:\02-Blades\17-Warehouse-Operators\Levantamento de Material em Armazém\Material Adicional\EXCEL\historico_base.xlsx")
    atualização_BD()
    atualizarfarol_PP()
    tempo = datetime.now() - data_hora_atual
    print("Finalizando em:", tempo)
    print('**************BD Relatório Status de Produção atualizado com sucesso!*************')

def atualizar_backup():
    # Obter a data e hora atuais
    data_hora_atual = datetime.now()
    # Formatar a data e hora em uma string
    nomearquivo = data_hora_atual.strftime("%d-%m-%y %H%M%S")
    # Exibir a data e hora formatadas
    print("Data e hora atuais formatadas:", nomearquivo)
    fechar_outlook()

    # Obter a data de inicio e fim do periodo de pesquisa
    data_inicio = datetime.now() - timedelta(days=35)
    data_fim = datetime.now() + timedelta(days=35)
    # Formatar a data e hora em uma string
    datainicio = data_inicio.strftime("%d%m%Y")
    datafim = data_fim.strftime("%d%m%Y")

    abrirrescheduling()
    sleep(5)
    exerescheduling()
    sleep(5)
    export_backupr()
    sleep(5)
    print('Backup realizado com sucesso - Rescheduling')
    sleep(5)

    abrirbackupcoois()
    sleep(5)
    execoois()
    sleep(5)
    clicarno()
    sleep(30)
    export_backupcoois()
    sleep(5)
    print('Backup realizado com sucesso - Coois (ordens planeadas/ordens convertidas)')
    
    sleep(5)
    abrirbackupzmb52()
    sleep(5)
    exezmb52()
    sleep(5)
    export_backupzmb52()
    sleep(5)
    print('Backup realizado com sucesso - ZMB52')
    sleep(5)

    abrirbackupzbom()
    sleep(5)
    exezbom_e115()
    sleep(5)
    export_menu()
    sleep(5)
    export_backupzbom_e115()
    sleep(5)
    print('Backup realizado com sucesso - ZBOM E115')
    sleep(5)

    exezbom_e175()
    sleep(5)
    export_menu()
    sleep(5)
    export_backupzbom_e175()
    sleep(5)
    print('Backup realizado com sucesso - ZBOM E175')
    sleep(5)
    
    exezbom_e103()
    sleep(5)
    export_menu()
    sleep(5)
    export_backupzbom_e103()
    sleep(5)
    print('Backup realizado com sucesso - ZBOM E103')
    sleep(5)

    abrirbackupp2p()
    sleep(5)
    exep2p()
    sleep(5)
    export_backup2p()
    sleep(5)
    print('Backup realizado com sucesso - Mapa responsabilidade')
    sleep(5)
    pasta_origem = r'O:\02-Blades\04-Production\04 - Production Planning\2025'
    pasta_destino = r'C:\Users\00082300\ENERCON\PT ROTO BCM Share - Priority\Planeamento da produção\2025'
    copiar_arquivos(pasta_origem, pasta_destino)
    tempo = datetime.now() - data_hora_atual
    print("Finalizando em:", tempo)

def abrirscm():
    # Abrir o aplicativo
    os.startfile(r"C:\Users\00082300\Downloads\Projeto 2\atalhos\scm.SAP")
    sleep(10)

    # Tentar encontrar a janela aberta
    windows = gw.getWindowsWithTitle('SCP')

    if windows:
        app_window = windows[0]
        app_window.maximize()
    else:
        print("Janela do aplicativo não encontrada")
def exescm():
    # Obter a data e hora atuais
    data_hora_atual = datetime.now()
    # Formatar a data e hora em uma string
    nomearquivo = data_hora_atual.strftime("%d-%m-%y %H%M%S")
    # Exibir a data e hora formatadas
    print("Data e hora atuais formatadas:", nomearquivo)

    # Obter a data de inicio e fim do periodo de pesquisa
    data_inicio = datetime.now()
    data_fim = datetime.now() + timedelta(days=35)
    # Formatar a data e hora em uma string
    datainicio = data_inicio.strftime("%d%m%Y")
    datafim = data_fim.strftime("%d%m%Y")

    pyautogui.hotkey('shift', 'f5')
    sleep(4)

    esperar_e_clicar('variantescm.png', "Confirmar Variavel", timeout=30, clicks=2)
    sleep(2)
    pyautogui.write(scm)

    esperar_e_clicar('conf_variavel.png', "Confirmar Variavel", timeout=30, clicks=2)
    esperar_e_clicar('datestart.png', "Confirmar Variavel", timeout=30, clicks=2)
    esperar_e_clicar('X_vemerlho.png', "Confirmar Variavel", timeout=30, clicks=1)
    imagem = 'X_vemerlho.png'
    sleep(3)
    pyautogui.write(datainicio)

    esperar_e_clicar('000.png', "Confirmar Variavel", timeout=30, clicks=1)
    pyautogui.hotkey('tab')
    pyautogui.write("*")
    
    pyautogui.hotkey('f8')
    sleep(60)
    print("Em execução a transação")
def export_scm():
    # Obter a data e hora atuais
    data_hora_atual = datetime.now()
    # Formatar a data e hora em uma string
    nomearquivo = data_hora_atual.strftime("%d-%m-%y %H%M%S")
    # Exibir a data e hora formatadas
    print("Data e hora atuais formatadas:", nomearquivo)

    pyautogui.hotkey('shift','f9')
    sleep(20)

    arquivo = r'O:\02-Blades\02-Process Engineering\9. Projetos\23. Controle SCM\BD_diario\SCM_' + nomearquivo
    pyautogui.write(arquivo)
    pyautogui.hotkey('enter')
    sleep(10)
    pyautogui.hotkey('alt','f4')
    sleep(10)

    esperar_e_clicar('sair_x.png', "Confirmar Variavel", timeout=30, clicks=2)
    esperar_e_clicar('yes_sair.png', "Confirmar Variavel", timeout=30, clicks=1)

    print("Export finalizado!")

def atualizarscm():
    # Obter a data e hora atuais
    data_hora_atual = datetime.now()
    # Formatar a data e hora em uma string
    nomearquivo = data_hora_atual.strftime("%d-%m-%y %H%M%S")
    # Exibir a data e hora formatadas
    print("Data e hora atuais formatadas:", nomearquivo)
    fechar_outlook()
    abrirscm()
    exescm()
    export_scm()
    tempo = datetime.now() - data_hora_atual
    print("Finalizando em:", tempo)

def backup_log():
    pasta_origem = r'\\srv-pt3\groups\02-Blades\05-Warehouse\ARMAZÉM\Fábrica Sul\3. Folhas para Carrinhos - Matéria Prima\4. Material por Peça E103'
    pasta_destino = r'C:\Users\00082300\ENERCON\PT ROTO BCM Share - Priority\Logistica - Armazem\IT-Breakdown MPP'
    copiar_arquivos(pasta_origem, pasta_destino)
    print("Backup realizado com sucesso - 4. Material por Peça E103 - Fábrica Sul")
    pasta_origem = r'\\srv-pt3\groups\02-Blades\05-Warehouse\ARMAZÉM\Fábrica Sul\3. Folhas para Carrinhos - Matéria Prima\5. Material por peça E175'
    pasta_destino = r'C:\Users\00082300\ENERCON\PT ROTO BCM Share - Priority\Logistica - Armazem\IT-Breakdown MPP'
    copiar_arquivos(pasta_origem, pasta_destino)
    print("Backup realizado com sucesso - 5. Material por peça E175 - Fábrica Sul")
    pasta_origem = r'\\srv-pt3\groups\02-Blades\05-Warehouse\ARMAZÉM\Fábrica Norte\2. MPP\MPP 175'
    pasta_destino = r'C:\Users\00082300\ENERCON\PT ROTO BCM Share - Priority\Logistica - Armazem\IT-Breakdown MPP'
    copiar_arquivos(pasta_origem, pasta_destino)
    print("Backup realizado com sucesso - MPP 175 - Fábrica Norte")
    pasta_origem = r'\\srv-pt3\groups\02-Blades\05-Warehouse\ARMAZÉM\Fábrica Norte\2. MPP\MPP103'
    pasta_destino = r'C:\Users\00082300\ENERCON\PT ROTO BCM Share - Priority\Logistica - Armazem\IT-Breakdown MPP'
    copiar_arquivos(pasta_origem, pasta_destino)
    print("Backup realizado com sucesso - MPP103 - Fábrica Norte")

