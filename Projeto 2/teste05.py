import openpyxl
from openpyxl import load_workbook
from datetime import datetime
import os
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.application import MIMEApplication
from email.mime.text import MIMEText
from tkinter import Tk, Label, Entry, Button, OptionMenu, StringVar

# Função para gerar o arquivo Excel e preenchê-lo
def gerar_excel(turno, tipo_material, sap, material, qtd, st_loc, num_peca, motivo, sap_op_armazen, sap_op_producao, nome_op_producao):
    # Carregar modelo
    wb = load_workbook('rF-158 Material Adicional_Armazem_Sul.xlsm', keep_vba=True)
    ws = wb.active

    # Preencher dados no Excel
    ws['B5'] = turno
    ws['D5'] = datetime.now().strftime("%d/%m/%Y %H:%M")
    ws['B7'] = tipo_material
    ws['B9'] = sap
    ws['C9'] = material
    ws['D9'] = qtd
    ws['E9'] = st_loc
    ws['F9'] = num_peca
    ws['G9'] = motivo
    ws['H9'] = sap_op_armazen
    ws['I9'] = sap_op_producao
    ws['J9'] = nome_op_producao

    # Salvar arquivo Excel
    nome_arquivo = f"Material_Adicional_{datetime.now().strftime('%Y%m%d%H%M%S')}.xlsm"
    wb.save(nome_arquivo)

    # Converter para PDF e retornar o caminho
    return nome_arquivo

# Função para enviar o arquivo por e-mail
def enviar_email(destinatario, arquivo):
    email_user = 'seuemail@gmail.com'
    email_password = 'suasenha'
    email_send = destinatario

    subject = 'Envio de Material Adicional'

    msg = MIMEMultipart()
    msg['From'] = email_user
    msg['To'] = email_send
    msg['Subject'] = subject

    body = 'Segue o material adicional solicitado.'
    msg.attach(MIMEText(body, 'plain'))

    # Anexar arquivo
    with open(arquivo, 'rb') as f:
        part = MIMEApplication(f.read(), Name=os.path.basename(arquivo))
        part['Content-Disposition'] = f'attachment; filename="{os.path.basename(arquivo)}"'
        msg.attach(part)

    # Configurar servidor SMTP e enviar e-mail
    server = smtplib.SMTP('smtp.gmail.com', 587)
    server.starttls()
    server.login(email_user, email_password)
    text = msg.as_string()
    server.sendmail(email_user, email_send, text)
    server.quit()

# Interface Gráfica para o usuário preencher os dados
def interface():
    def cadastrar():
        turno = turno_var.get()
        tipo_material = tipo_material_var.get()
        sap = sap_entry.get()
        material = material_entry.get()
        qtd = qtd_entry.get()
        st_loc = st_loc_entry.get()
        num_peca = num_peca_entry.get()
        motivo = motivo_entry.get()
        sap_op_armazen = sap_op_armazen_entry.get()
        sap_op_producao = sap_op_producao_entry.get()
        nome_op_producao = nome_op_producao_entry.get()

        arquivo = gerar_excel(turno, tipo_material, sap, material, qtd, st_loc, num_peca, motivo, sap_op_armazen, sap_op_producao, nome_op_producao)
        enviar_email('destinatario@example.com', arquivo)

        status_label.config(text="Arquivo gerado e enviado com sucesso!")

    # Janela principal
    root = Tk()
    root.title("Cadastro de Material Adicional")

    Label(root, text="Turno").grid(row=0, column=0)
    turno_var = StringVar(root)
    turno_entry = Entry(root, textvariable=turno_var).grid(row=0, column=1)

    Label(root, text="Tipo de Material").grid(row=1, column=0)
    tipo_material_var = StringVar(root)
    tipo_material_menu = OptionMenu(root, tipo_material_var, "Short Rolls", "Material Extra Adicional", "Manutenção de Moldes", "Reparações")
    tipo_material_menu.grid(row=1, column=1)

    Label(root, text="SAP").grid(row=2, column=0)
    sap_entry = Entry(root).grid(row=2, column=1)

    Label(root, text="Material").grid(row=3, column=0)
    material_entry = Entry(root).grid(row=3, column=1)

    Label(root, text="Qt. Requesitada").grid(row=4, column=0)
    qtd_entry = Entry(root).grid(row=4, column=1)

    Label(root, text="St Loc.").grid(row=5, column=0)
    st_loc_entry = Entry(root).grid(row=5, column=1)

    Label(root, text="Número da Peça").grid(row=6, column=0)
    num_peca_entry = Entry(root).grid(row=6, column=1)

    Label(root, text="Motivo").grid(row=7, column=0)
    motivo_entry = Entry(root).grid(row=7, column=1)

    Label(root, text="SAP Operador Armazém").grid(row=8, column=0)
    sap_op_armazen_entry = Entry(root).grid(row=8, column=1)

    Label(root, text="SAP Operador Produção").grid(row=9, column=0)
    sap_op_producao_entry = Entry(root).grid(row=9, column=1)

    Label(root, text="Nome do Operador de Produção").grid(row=10, column=0)
    nome_op_producao_entry = Entry(root).grid(row=10, column=1)

    Button(root, text="Cadastrar", command=cadastrar).grid(row=11, column=0, columnspan=2)

    status_label = Label(root, text="")
    status_label.grid(row=12, column=0, columnspan=2)

    root.mainloop()

# Executa a interface
interface()
